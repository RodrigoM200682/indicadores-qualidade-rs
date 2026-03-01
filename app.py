# ============================================================
# INDICADORES DE QUALIDADE RS — QualiEx (versão completa)
#
# ATUALIZAÇÃO (solicitada agora):
# - Se Ano = (Todos): gráfico principal mostra TOTAL por MÊS/ANO no eixo X (ex.: Jan/2025)
#   * Clique na barra: abre lista detalhada daquele mês/ano (por responsável análise)
# - Se Ano específico:
#   * Se Evolução = Mensal: gráfico por mês (Jan..Dez)
#       - Clique na barra do mês: abre POPUP com gráfico por semana do mês (1..6)
#           - Clique na barra da semana: abre lista detalhada daquela semana (por responsável)
#   * Se Evolução = Semanal: mantém o comportamento atual (semana 1..52 ou semana do mês se mês selecionado)
#
# Mantém:
# - Filtros laterais completos (inclui Categoria se existir)
# - Popup agrupado por Responsável da análise de causa + exportar recorte
# - Resumo Excel bonito (DASHBOARD com 4 gráficos um abaixo do outro + tabelas)
#
# Dependências:
#   py -m pip install pandas matplotlib openpyxl
# ============================================================

import os
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.ticker import MaxNLocator
from matplotlib.figure import Figure

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart
from openpyxl.chart.reference import Reference


# =========================
# Regras de cores do gráfico
# =========================
LIMIAR_SEMANAL = 2  # <=2 verde, >2 vermelho
LIMIAR_MENSAL = 8   # <=8 verde, >8 vermelho

# =========================
# Base Reclamações
# =========================
DEFAULT_EXCEL_RECLAMACOES = "Consultas_RNC.xlsx"
DEFAULT_SHEET = "Sheet1"

COL_CODIGO = "Código"
COL_TITULO = "Título"
COL_STATUS = "Status"
COL_DATA = "Data de emissão"
COL_MOTIVO = "Motivo Reclamação"
COL_TURNO = "Turno/Horário"

COL_RESP_OCORRENCIA = "Responsável"
COL_RESP_ANALISE = "Responsável da análise de causa"
COL_SITUACAO = "Situação"  # ATRASADA / NO PRAZO

FILTROS_COLS_RECLAMACOES = [
    "Status",
    "Categoria",
    "Cliente",
    "Motivo Reclamação",
    "Responsável",
    "Responsável da análise de causa",
    "Turno/Horário",
    "Embalagem",
    COL_SITUACAO,
]

DATE_FMT_BR = "%d/%m/%Y"
MESES_ABREV = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
}
INV_MESES_ABREV = {v: k for k, v in MESES_ABREV.items()}


# ============================================================
# Utilitários
# ============================================================
def mes_abrev(n: int) -> str:
    return MESES_ABREV.get(int(n), str(n).zfill(2))


def br_date_str(dt) -> str:
    if pd.isna(dt):
        return ""
    try:
        return pd.to_datetime(dt).strftime(DATE_FMT_BR)
    except Exception:
        return ""


def normalizar_situacao(x: str) -> str:
    s = str(x).strip().upper()
    if s in ("ATRASADA", "ATRASADO"):
        return "ATRASADA"
    if s in ("NO PRAZO", "NOPRAZO", "NO_PRAZO"):
        return "NO PRAZO"
    if s in ("", "NAN", "NONE"):
        return ""
    return s


def normalizar_status(x: str) -> str:
    return str(x).strip().upper()


def is_reprovada(status: str) -> bool:
    return "REPROV" in normalizar_status(status)


def is_associada(status: str) -> bool:
    return "ASSOCI" in normalizar_status(status)


def semana_1a52(series_datetime: pd.Series) -> pd.Series:
    w = series_datetime.dt.isocalendar().week.astype(int)
    return w.clip(upper=52)


def semana_no_mes(series_datetime: pd.Series) -> pd.Series:
    return ((series_datetime.dt.day - 1) // 7 + 1).astype(int).clip(upper=6)


def ano_padrao_para_relatorios(df: pd.DataFrame, col_data: str) -> int:
    anos = sorted(df[col_data].dt.year.dropna().unique().tolist())
    return int(anos[-1]) if anos else int(pd.Timestamp.today().year)


def contar_top(df, col: str, top_n=10, vazio="SEM VALOR"):
    if df.empty or col not in df.columns:
        return pd.Series(dtype=int)
    s = df[col].fillna("").replace("", vazio).value_counts()
    if len(s) <= top_n:
        return s
    top = s.iloc[:top_n].copy()
    top.loc["OUTROS"] = int(s.iloc[top_n:].sum())
    return top


def carregar_dados_reclamacoes(caminho, aba):
    df = pd.read_excel(caminho, sheet_name=aba)

    obrig = [COL_CODIGO, COL_TITULO, COL_STATUS, COL_DATA, COL_MOTIVO]
    for c in obrig:
        if c not in df.columns:
            raise ValueError(f"Não encontrei a coluna obrigatória '{c}' na planilha.")

    df = df.copy()
    df[COL_DATA] = pd.to_datetime(df[COL_DATA], errors="coerce", dayfirst=True)
    df = df.dropna(subset=[COL_DATA])

    for c in df.columns:
        if df[c].dtype == "object":
            df[c] = df[c].astype(str).str.strip().replace("nan", "")

    if COL_SITUACAO in df.columns:
        df[COL_SITUACAO] = df[COL_SITUACAO].apply(normalizar_situacao)

    return df


# ============================================================
# Excel helpers (relatório bonito)
# ============================================================
def _excel_theme():
    return {
        "title_fill": PatternFill("solid", fgColor="1F4E79"),
        "sub_fill": PatternFill("solid", fgColor="D9E1F2"),
        "hdr_fill": PatternFill("solid", fgColor="2F5597"),
        "hdr_font": Font(bold=True, color="FFFFFF"),
        "bold": Font(bold=True),
        "thin": Side(style="thin", color="A6A6A6"),
        "kpi_fill": PatternFill("solid", fgColor="FFF2CC"),
    }


def _xl_col(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def _apply_border(ws, cell_range: str):
    th = _excel_theme()["thin"]
    border = Border(left=th, right=th, top=th, bottom=th)
    for row in ws[cell_range]:
        for c in row:
            c.border = border


def _merge_title(ws, cell_range: str, text: str):
    theme = _excel_theme()
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.fill = theme["title_fill"]
    c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _add_table(ws, start_row, start_col, df: pd.DataFrame, table_name: str, style="TableStyleMedium9"):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, v in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=v)

    end_row = start_row + len(df)
    end_col = start_col + df.shape[1] - 1

    ref = f"{_xl_col(start_col)}{start_row}:{_xl_col(end_col)}{end_row}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tab)

    theme = _excel_theme()
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=c)
        cell.fill = theme["hdr_fill"]
        cell.font = theme["hdr_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col).coordinate
    ws.auto_filter.ref = ref
    _apply_border(ws, ref)
    return (start_row, start_col, end_row, end_col, ref)


def _chart_set_integer_y(chart: BarChart):
    try:
        chart.y_axis.majorUnit = 1
    except Exception:
        pass
    try:
        chart.y_axis.crosses = "autoZero"
    except Exception:
        pass


def _chart_set_x_45(chart: BarChart):
    try:
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.txPr = None
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickMarkSkip = 1
        chart.x_axis.textRotation = 45
    except Exception:
        pass


def _add_bar_chart(ws, title, cat_col, val_col, start_row, end_row, anchor_cell,
                   y_title="Ocorrências", x_title="", rotate_x_45=False):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.legend = None
    chart.height = 9.0
    chart.width = 26.0

    data = Reference(ws, min_col=val_col, min_row=start_row, max_row=end_row)
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    _chart_set_integer_y(chart)
    if rotate_x_45:
        _chart_set_x_45(chart)

    ws.add_chart(chart, anchor_cell)


def _safe_sheetname(name: str) -> str:
    bad = ['\\', '/', '*', '?', ':', '[', ']']
    for b in bad:
        name = name.replace(b, " ")
    name = name.strip()
    return name[:31] if len(name) > 31 else name


# ============================================================
# Sidebar scroll
# ============================================================
def criar_sidebar_scroll(parent, width=320):
    canvas = tk.Canvas(parent, width=width, highlightthickness=0)
    sb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=sb.set)

    sb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    inner = ttk.Frame(canvas)
    win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

    def _on_inner_configure(_event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def _on_canvas_configure(event):
        canvas.itemconfigure(win_id, width=event.width)

    inner.bind("<Configure>", _on_inner_configure)
    canvas.bind("<Configure>", _on_canvas_configure)

    def _on_mousewheel(event):
        if event.delta:
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_linux_up(_event):
        canvas.yview_scroll(-3, "units")

    def _on_linux_down(_event):
        canvas.yview_scroll(3, "units")

    def _bind_wheel(_event):
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", _on_linux_up)
        canvas.bind_all("<Button-5>", _on_linux_down)

    def _unbind_wheel(_event):
        canvas.unbind_all("<MouseWheel>")
        canvas.unbind_all("<Button-4>")
        canvas.unbind_all("<Button-5>")

    canvas.bind("<Enter>", _bind_wheel)
    canvas.bind("<Leave>", _unbind_wheel)

    return canvas, inner


# ============================================================
# ChecklistFilter
# ============================================================
class ChecklistFilter(ttk.LabelFrame):
    def __init__(self, master, title, on_change, height=90):
        super().__init__(master, text=title, padding=4)
        self.on_change = on_change
        self.vars = {}
        self.values_cache = []

        top = ttk.Frame(self)
        top.pack(fill="x", pady=(0, 3))
        ttk.Button(top, text="Todos", command=self.select_all, width=7).pack(side="left")
        ttk.Button(top, text="Nenhum", command=self.select_none, width=7).pack(side="left", padx=4)

        self.canvas = tk.Canvas(self, height=height, highlightthickness=0)
        self.scroll = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview, width=12)
        self.inner = ttk.Frame(self.canvas)

        self.inner.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self._win_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scroll.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scroll.pack(side="right", fill="y")
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfigure(self._win_id, width=e.width))

        def _on_mousewheel(event):
            if event.delta:
                self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _on_linux_up(_event):
            self.canvas.yview_scroll(-3, "units")

        def _on_linux_down(_event):
            self.canvas.yview_scroll(3, "units")

        def _bind_wheel(_event):
            self.canvas.bind_all("<MouseWheel>", _on_mousewheel)
            self.canvas.bind_all("<Button-4>", _on_linux_up)
            self.canvas.bind_all("<Button-5>", _on_linux_down)

        def _unbind_wheel(_event):
            self.canvas.unbind_all("<MouseWheel>")
            self.canvas.unbind_all("<Button-4>")
            self.canvas.unbind_all("<Button-5>")

        self.canvas.bind("<Enter>", _bind_wheel)
        self.canvas.bind("<Leave>", _unbind_wheel)

    def set_values_preserve(self, values):
        values = list(values)
        if values == self.values_cache:
            return

        selected_now = set(self.get_selected())

        for w in self.inner.winfo_children():
            w.destroy()
        self.vars.clear()
        self.values_cache = values

        for v in values:
            var = tk.BooleanVar(value=True)
            if selected_now:
                var.set(v in selected_now)
            self.vars[v] = var
            cb = ttk.Checkbutton(self.inner, text=v, variable=var, command=self.on_change)
            cb.pack(anchor="w", pady=0)

        if values and not self.get_selected():
            self.select_all()

    def get_selected(self):
        return [v for v, var in self.vars.items() if var.get()]

    def select_all(self):
        for var in self.vars.values():
            var.set(True)
        self.on_change()

    def select_none(self):
        for var in self.vars.values():
            var.set(False)
        self.on_change()


# ============================================================
# Aba Reclamações
# ============================================================
class ReclamacoesClienteFrame(ttk.Frame):
    def __init__(self, master):
        super().__init__(master)

        self.caminho_excel = tk.StringVar(value=self._excel_padrao())

        self.sel_ano = tk.StringVar(value="(Todos)")
        self.sel_mes = tk.StringVar(value="(Todos)")
        self.sel_resp_occ = tk.StringVar(value="(Todos)")
        self.granularidade = tk.StringVar(value="Semanal")

        self.filters = {}
        self.df_base = pd.DataFrame()
        self.df_filtrado = pd.DataFrame()
        self._pick_map = {}

        self._montar_ui()
        self._atualizar()

    def _excel_padrao(self):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_dir, DEFAULT_EXCEL_RECLAMACOES)

    def _montar_ui(self):
        try:
            style = ttk.Style()
            style.configure("TLabel", font=("Segoe UI", 9))
            style.configure("TButton", font=("Segoe UI", 9))
            style.configure("TCheckbutton", font=("Segoe UI", 9))
            style.configure("TLabelframe.Label", font=("Segoe UI", 9, "bold"))
        except Exception:
            pass

        topo = ttk.Frame(self, padding=10)
        topo.pack(fill="x")

        ttk.Label(topo, text="Arquivo Excel:").grid(row=0, column=0, sticky="w")
        ttk.Entry(topo, textvariable=self.caminho_excel, width=80).grid(row=0, column=1, padx=6)
        ttk.Button(topo, text="Procurar...", command=self._procurar_excel).grid(row=0, column=2, padx=(0, 10))
        ttk.Button(topo, text="Atualizar", command=self._atualizar).grid(row=0, column=3)
        topo.grid_columnconfigure(1, weight=1)

        cfg = ttk.Frame(self, padding=(10, 0, 10, 10))
        cfg.pack(fill="x")

        ttk.Label(cfg, text="Ano:").grid(row=0, column=0, sticky="e")
        self.cb_ano = ttk.Combobox(cfg, textvariable=self.sel_ano, state="readonly", width=10)
        self.cb_ano.grid(row=0, column=1, padx=6, sticky="w")

        ttk.Label(cfg, text="Mês:").grid(row=0, column=2, padx=(10, 0), sticky="e")
        self.cb_mes = ttk.Combobox(cfg, textvariable=self.sel_mes, state="readonly", width=10)
        self.cb_mes.grid(row=0, column=3, padx=6, sticky="w")

        ttk.Label(cfg, text="Resp. ocorrência:").grid(row=0, column=4, padx=(10, 0), sticky="e")
        self.cb_resp = ttk.Combobox(cfg, textvariable=self.sel_resp_occ, state="readonly", width=22)
        self.cb_resp.grid(row=0, column=5, padx=6, sticky="w")

        ttk.Label(cfg, text="Evolução:").grid(row=0, column=6, padx=(10, 0), sticky="e")
        self.cb_gran = ttk.Combobox(cfg, textvariable=self.granularidade, state="readonly",
                                    values=["Semanal", "Mensal"], width=10)
        self.cb_gran.grid(row=0, column=7, padx=6, sticky="w")

        ttk.Button(cfg, text="Aplicar", command=self._render).grid(row=0, column=8, padx=(14, 6))
        ttk.Button(cfg, text="Gerar Resumo Excel (bonito)", command=self._gerar_resumo_excel_bonito_do_filtro)\
            .grid(row=0, column=9, padx=(6, 0))

        self.paned = ttk.Panedwindow(self, orient="horizontal")
        self.paned.pack(fill="both", expand=True, padx=10, pady=10)

        self.filtros_frame = ttk.LabelFrame(self.paned, text="Filtros por marcar", padding=8)
        self.painel_frame = ttk.Frame(self.paned)

        self.paned.add(self.filtros_frame, weight=1)
        self.paned.add(self.painel_frame, weight=4)

        _, inner = criar_sidebar_scroll(self.filtros_frame, width=320)

        for col in FILTROS_COLS_RECLAMACOES:
            if str(col).strip().lower() in ("unidade", "local"):
                continue
            f = ChecklistFilter(inner, col, on_change=self._render, height=90)
            f.pack(fill="x", pady=4)
            self.filters[col] = f

        self.fig = Figure(figsize=(11.4, 6.2), dpi=110)
        gs = self.fig.add_gridspec(nrows=2, ncols=1, height_ratios=[3, 2])
        self.ax1 = self.fig.add_subplot(gs[0, 0])
        self.ax2 = self.fig.add_subplot(gs[1, 0])

        self.canvas_plot = FigureCanvasTkAgg(self.fig, master=self.painel_frame)
        self.canvas_plot.get_tk_widget().pack(fill="x", expand=False, pady=(0, 10))
        self.canvas_plot.mpl_connect("pick_event", self._on_pick)

        tabela_frame = ttk.LabelFrame(self.painel_frame, text="Ocorrências (filtro atual)", padding=8)
        tabela_frame.pack(fill="both", expand=True)

        cols = ("Código", "Título", "Status", "Data de emissão", "Turno/Horário")
        self.tree = ttk.Treeview(tabela_frame, columns=cols, show="headings", height=18)

        for col in cols:
            self.tree.heading(col, text=col)
            if col == "Título":
                self.tree.column(col, width=720, anchor="w")
            else:
                self.tree.column(col, width=170, anchor="w")

        self.tree.tag_configure("reprovada", foreground="blue")
        self.tree.tag_configure("associada", foreground="#A9A9A9")
        self.tree.tag_configure("normal", foreground="black")

        self.tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(tabela_frame, orient="vertical", command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=sb.set)

        self.lbl_info = ttk.Label(self, text="", padding=10)
        self.lbl_info.pack(fill="x")

    def _procurar_excel(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o Excel",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls")]
        )
        if arquivo:
            self.caminho_excel.set(arquivo)
            self._atualizar()

    def _atualizar(self):
        try:
            self.df_base = carregar_dados_reclamacoes(self.caminho_excel.get(), DEFAULT_SHEET)
            self._popular_filtros()
            self._render()
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def _popular_filtros(self):
        df = self.df_base

        anos = sorted(df[COL_DATA].dt.year.dropna().unique().tolist())
        self.cb_ano["values"] = ["(Todos)"] + [str(a) for a in anos]
        if self.sel_ano.get() not in self.cb_ano["values"]:
            self.sel_ano.set("(Todos)")

        self.cb_mes["values"] = ["(Todos)"] + [MESES_ABREV[m] for m in range(1, 13)]
        if self.sel_mes.get() not in self.cb_mes["values"]:
            self.sel_mes.set("(Todos)")

        if COL_RESP_OCORRENCIA in df.columns:
            resp_vals = sorted(df[COL_RESP_OCORRENCIA].dropna().astype(str).replace("nan", "").unique().tolist())
            resp_vals = [v for v in resp_vals if v != ""]
        else:
            resp_vals = []
        self.cb_resp["values"] = ["(Todos)"] + resp_vals
        if self.sel_resp_occ.get() not in self.cb_resp["values"]:
            self.sel_resp_occ.set("(Todos)")

        for col, widget in self.filters.items():
            if col in df.columns:
                vals = sorted(df[col].dropna().astype(str).replace("nan", "").unique().tolist())
                vals = [v for v in vals if v != ""]
            else:
                vals = []
            widget.set_values_preserve(vals)

        self._aplicar_regra_granularidade()

    def _aplicar_regra_granularidade(self):
        """
        NOVO COMPORTAMENTO:
        - Ano = (Todos): evolução travada em "Mensal" (pois será Mês/Ano no gráfico)
        - Ano específico: mantém Semanal e Mensal
        """
        if self.sel_ano.get() == "(Todos)":
            self.granularidade.set("Mensal")
            self.cb_gran["values"] = ["Mensal"]
        else:
            if self.granularidade.get() not in ("Semanal", "Mensal"):
                self.granularidade.set("Semanal")
            self.cb_gran["values"] = ["Semanal", "Mensal"]

    def _mes_to_num(self, mes_str):
        if mes_str == "(Todos)":
            return None
        return INV_MESES_ABREV.get(mes_str)

    def _aplicar_filtros(self):
        df_f = self.df_base.copy()
        self._aplicar_regra_granularidade()

        ano = self.sel_ano.get()
        mes_num = self._mes_to_num(self.sel_mes.get())

        if ano != "(Todos)":
            df_f = df_f[df_f[COL_DATA].dt.year == int(ano)]
        if mes_num is not None:
            df_f = df_f[df_f[COL_DATA].dt.month == int(mes_num)]

        if self.sel_resp_occ.get() != "(Todos)" and COL_RESP_OCORRENCIA in df_f.columns:
            df_f = df_f[df_f[COL_RESP_OCORRENCIA].astype(str) == self.sel_resp_occ.get()]

        for col, widget in self.filters.items():
            if col not in df_f.columns:
                continue
            selecionados = widget.get_selected()
            if not selecionados:
                return df_f.iloc[0:0]
            df_f = df_f[df_f[col].astype(str).isin(selecionados)]

        return df_f

    def _titulo_do_filtro_atual(self) -> str:
        partes = ["Reclamações — Filtro atual"]
        if self.sel_ano.get() != "(Todos)":
            partes.append(f"Ano {self.sel_ano.get()}")
        if self.sel_mes.get() != "(Todos)":
            partes.append(f"Mês {self.sel_mes.get()}")
        if self.sel_resp_occ.get() != "(Todos)":
            partes.append(f"Resp ocorrência: {self.sel_resp_occ.get()}")
        partes.append(f"Evolução: {self.granularidade.get()}")
        return " | ".join(partes)

    # =========================================================
    # RESUMO (4 gráficos um abaixo do outro + tabela de totais)
    # =========================================================
    def _gerar_resumo_excel_bonito_do_filtro(self):
        try:
            if self.df_base is None or self.df_base.empty:
                messagebox.showinfo("Resumo", "Base vazia. Carregue o Excel primeiro.")
                return

            self.df_filtrado = self._aplicar_filtros()
            dff = self.df_filtrado.copy()
            if dff.empty:
                messagebox.showinfo("Resumo", "Não há registros no filtro atual para gerar o resumo.")
                return

            titulo_filtro = self._titulo_do_filtro_atual()

            if self.sel_ano.get() != "(Todos)":
                ano_ref = int(self.sel_ano.get())
            else:
                ano_ref = ano_padrao_para_relatorios(self.df_base, COL_DATA)

            df_ano = self.df_base.copy()
            df_ano = df_ano[df_ano[COL_DATA].dt.year == ano_ref].copy()

            resp_ano = (
                df_ano[COL_RESP_ANALISE].fillna("").replace("", "SEM RESPONSÁVEL")
                if COL_RESP_ANALISE in df_ano.columns else pd.Series(["SEM RESPONSÁVEL"] * len(df_ano))
            )
            sit_ano = (
                df_ano[COL_SITUACAO].fillna("").apply(normalizar_situacao)
                if COL_SITUACAO in df_ano.columns else pd.Series([""] * len(df_ano))
            )
            atrasadas_por_resp_ano = (
                pd.DataFrame({"Responsável (análise)": resp_ano, "Situação": sit_ano})
                .query("Situação == 'ATRASADA'")["Responsável (análise)"].value_counts()
            )

            resp_rec = (
                dff[COL_RESP_ANALISE].fillna("").replace("", "SEM RESPONSÁVEL")
                if COL_RESP_ANALISE in dff.columns else pd.Series(["SEM RESPONSÁVEL"] * len(dff))
            )
            sit_rec = (
                dff[COL_SITUACAO].fillna("").apply(normalizar_situacao)
                if COL_SITUACAO in dff.columns else pd.Series([""] * len(dff))
            )

            total_rec = int(len(dff))
            total_atras_rec = int((sit_rec == "ATRASADA").sum()) if len(sit_rec) else 0

            p_ini = br_date_str(dff[COL_DATA].min()) if COL_DATA in dff.columns else "-"
            p_fim = br_date_str(dff[COL_DATA].max()) if COL_DATA in dff.columns else "-"

            dff_mes = dff.copy()
            dff_mes["MesNum"] = dff_mes[COL_DATA].dt.month.astype(int)
            g_mes = dff_mes.groupby("MesNum").size().reindex(range(1, 13), fill_value=0)
            df_mes = pd.DataFrame({
                "Mês": [MESES_ABREV[m] for m in range(1, 13)],
                "Ocorrências": g_mes.values.astype(int)
            })

            ocorr_por_resp_rec = resp_rec.value_counts()
            df_resp = pd.DataFrame({
                "Responsável (análise)": ocorr_por_resp_rec.index.astype(str),
                "Ocorrências": ocorr_por_resp_rec.values.astype(int)
            })
            if df_resp.empty:
                df_resp = pd.DataFrame({"Responsável (análise)": ["SEM DADOS"], "Ocorrências": [0]})

            df_atras_ano = pd.DataFrame({
                "Responsável (análise)": atrasadas_por_resp_ano.index.astype(str),
                "Atrasadas (ano todo)": atrasadas_por_resp_ano.values.astype(int)
            })
            if df_atras_ano.empty:
                df_atras_ano = pd.DataFrame({"Responsável (análise)": ["SEM DADOS"], "Atrasadas (ano todo)": [0]})

            top_mot = (
                dff[COL_MOTIVO].fillna("").replace("", "SEM MOTIVO").value_counts().head(12)
                if COL_MOTIVO in dff.columns else pd.Series(dtype=int)
            )
            df_mot = pd.DataFrame({
                "Motivo": top_mot.index.astype(str),
                "Ocorrências": top_mot.values.astype(int)
            })
            if df_mot.empty:
                df_mot = pd.DataFrame({"Motivo": ["SEM DADOS"], "Ocorrências": [0]})

            nome_padrao = f"Resumo_Filtro_{pd.Timestamp.today().strftime('%d-%m-%Y')}.xlsx"
            caminho = filedialog.asksaveasfilename(
                title="Salvar resumo em Excel",
                defaultextension=".xlsx",
                initialfile=nome_padrao,
                filetypes=[("Excel", "*.xlsx")]
            )
            if not caminho:
                return

            wb = Workbook()
            theme = _excel_theme()

            ws = wb.active
            ws.title = "DASHBOARD"
            ws.sheet_view.showGridLines = False
            _set_col_widths(ws, {
                "A": 2, "B": 34, "C": 16, "D": 16, "E": 16, "F": 16,
                "G": 16, "H": 16, "I": 16, "J": 16, "K": 16, "L": 16
            })

            _merge_title(ws, "B2:L2", "RESUMO DE OCORRÊNCIAS — 4 GRÁFICOS")
            ws.row_dimensions[2].height = 26

            ws["B4"] = "Filtro:"
            ws["B4"].font = Font(bold=True)
            ws["C4"] = titulo_filtro
            ws.merge_cells("C4:L4")
            ws["C4"].alignment = Alignment(wrap_text=True, vertical="top")

            ws["B6"] = "Período (recorte):"
            ws["B6"].font = Font(bold=True)
            ws["C6"] = f"{p_ini} a {p_fim}"

            ws["E6"] = "Ano ref (atrasadas ano todo):"
            ws["E6"].font = Font(bold=True)
            ws["G6"] = ano_ref

            for rng in ("B8:F12",):
                for row in ws[rng]:
                    for c in row:
                        c.fill = theme["kpi_fill"]
                        c.alignment = Alignment(vertical="center", wrap_text=True)
            _apply_border(ws, "B8:F12")

            ws["B8"] = "TOTAIS (FILTRO ATUAL)"
            ws["B8"].font = Font(bold=True, size=12)
            ws.merge_cells("B8:F8")

            ws["B10"] = "Total de ocorrências"
            ws["B10"].font = Font(bold=True)
            ws["C10"] = total_rec
            ws["C10"].font = Font(bold=True, size=14)

            ws["D10"] = "Ocorrências em atraso"
            ws["D10"].font = Font(bold=True)
            ws["E10"] = total_atras_rec
            ws["E10"].font = Font(bold=True, size=14)

            ws["B12"] = "Obs.: tabela completa por responsável está abaixo (seção Dados)."
            ws.merge_cells("B12:F12")

            row_data = 14
            ws["B14"] = "DADOS (não editar)"; ws["B14"].font = Font(bold=True, size=11)

            start1 = row_data + 2
            ws["B16"] = "1) Ocorrências por mês (filtro)"; ws["B16"].font = Font(bold=True)
            t1 = _add_table(ws, start1, 2, df_mes, table_name="T_MES", style="TableStyleMedium9")
            r1s, c1s, r1e, c1e, _ = t1

            g1_anchor = f"B{r1e + 2}"
            _add_bar_chart(
                ws, "Ocorrências por mês (filtro)", cat_col=2, val_col=3,
                start_row=r1s, end_row=r1e, anchor_cell=g1_anchor,
                y_title="Ocorrências", rotate_x_45=False
            )

            start2 = r1e + 20
            ws[f"B{start2-1}"] = "2) Ocorrências por responsável (análise) — filtro"; ws[f"B{start2-1}"].font = Font(bold=True)
            t2 = _add_table(ws, start2, 2, df_resp, table_name="T_RESP", style="TableStyleMedium9")
            r2s, c2s, r2e, c2e, _ = t2

            g2_anchor = f"B{r2e + 2}"
            _add_bar_chart(
                ws, "Ocorrências por responsável (análise) — filtro",
                cat_col=2, val_col=3,
                start_row=r2s, end_row=r2e, anchor_cell=g2_anchor,
                y_title="Ocorrências", rotate_x_45=True
            )

            start3 = r2e + 20
            ws[f"B{start3-1}"] = f"3) Atrasadas por responsável (análise) — ano {ano_ref} (ano todo)"; ws[f"B{start3-1}"].font = Font(bold=True)
            t3 = _add_table(ws, start3, 2, df_atras_ano, table_name="T_ATRAS_ANO", style="TableStyleMedium7")
            r3s, c3s, r3e, c3e, _ = t3

            g3_anchor = f"B{r3e + 2}"
            _add_bar_chart(
                ws, f"Atrasadas por responsável (análise) — ano {ano_ref}",
                cat_col=2, val_col=3,
                start_row=r3s, end_row=r3e, anchor_cell=g3_anchor,
                y_title="Atrasadas", rotate_x_45=True
            )

            try:
                rng = f"C{r3s+1}:C{r3e}"
                ws.conditional_formatting.add(
                    rng, CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE"))
                )
            except Exception:
                pass

            start4 = r3e + 20
            ws[f"B{start4-1}"] = "4) Ocorrências por motivo (Top 12) — filtro"; ws[f"B{start4-1}"].font = Font(bold=True)
            t4 = _add_table(ws, start4, 2, df_mot, table_name="T_MOT", style="TableStyleMedium9")
            r4s, c4s, r4e, c4e, _ = t4

            g4_anchor = f"B{r4e + 2}"
            _add_bar_chart(
                ws, "Ocorrências por motivo (Top 12) — filtro",
                cat_col=2, val_col=3,
                start_row=r4s, end_row=r4e, anchor_cell=g4_anchor,
                y_title="Ocorrências", rotate_x_45=True
            )

            ws.row_dimensions[4].height = 30

            ws2 = wb.create_sheet("RECORTE")
            ws2.sheet_view.showGridLines = True
            _merge_title(ws2, "A1:H1", "LISTA DE OCORRÊNCIAS — FILTRO ATUAL")
            ws2.row_dimensions[1].height = 26

            cols_doc = [COL_CODIGO, COL_TITULO, COL_STATUS, COL_DATA, COL_MOTIVO, COL_RESP_ANALISE, COL_SITUACAO]
            cols_doc = [c for c in cols_doc if c in dff.columns]

            dff_out = dff.copy()
            if COL_DATA in dff_out.columns:
                dff_out = dff_out.sort_values(COL_DATA, ascending=False).copy()
                dff_out[COL_DATA] = dff_out[COL_DATA].apply(br_date_str)

            if cols_doc:
                dff_out = dff_out[cols_doc].copy()

            start_row = 3
            _add_table(ws2, start_row, 1, dff_out, table_name="T_RECORTE", style="TableStyleMedium9")

            ws2.column_dimensions["A"].width = 14
            ws2.column_dimensions["B"].width = 70
            ws2.column_dimensions["C"].width = 16
            ws2.column_dimensions["D"].width = 16
            ws2.column_dimensions["E"].width = 26
            ws2.column_dimensions["F"].width = 30
            ws2.column_dimensions["G"].width = 14

            if COL_SITUACAO in cols_doc:
                sit_col_idx = cols_doc.index(COL_SITUACAO) + 1
                sit_col_letter = _xl_col(sit_col_idx)
                data_end = start_row + len(dff_out)
                rng = f"{sit_col_letter}{start_row+1}:{sit_col_letter}{data_end}"
                ws2.conditional_formatting.add(
                    rng, CellIsRule(operator="equal", formula=['"ATRASADA"'], fill=PatternFill("solid", fgColor="FFC7CE"))
                )
                ws2.conditional_formatting.add(
                    rng, CellIsRule(operator="equal", formula=['"NO PRAZO"'], fill=PatternFill("solid", fgColor="C6EFCE"))
                )

            wb.save(caminho)
            messagebox.showinfo("Resumo", f"Resumo Excel gerado com sucesso:\n{caminho}")

        except Exception as e:
            messagebox.showerror("Erro (Resumo Excel)", str(e))

    # =========================================================
    # Render principal
    # =========================================================
    def _render(self):
        if self.df_base is None or self.df_base.empty:
            return

        self.df_filtrado = self._aplicar_filtros()

        # -------- GRÁFICO PRINCIPAL (ax1) --------
        self.ax1.clear()
        self._pick_map = {}
        dfp = self.df_filtrado.copy()

        # NOVO: Ano = Todos => MÊS/ANO no eixo X
        if self.sel_ano.get() == "(Todos)":
            if dfp.empty:
                self.ax1.set_title("Evolução — Ano = Todos (sem dados no filtro)")
                self.ax1.text(0.5, 0.5, "Sem dados no filtro", ha="center", va="center")
            else:
                dfp["Ano"] = dfp[COL_DATA].dt.year.astype(int)
                dfp["Mes"] = dfp[COL_DATA].dt.month.astype(int)

                g = (
                    dfp.groupby(["Ano", "Mes"], as_index=False)
                    .size()
                    .rename(columns={"size": "Ocorrencias"})
                    .sort_values(["Ano", "Mes"])
                    .reset_index(drop=True)
                )
                g["Label"] = g.apply(lambda r: f"{MESES_ABREV[int(r['Mes'])]}/{int(r['Ano'])}", axis=1)

                x = list(range(len(g)))
                y = g["Ocorrencias"].astype(int).tolist()
                labels = g["Label"].astype(str).tolist()

                colors = ["green" if v <= LIMIAR_MENSAL else "red" for v in y]
                bars = self.ax1.bar(x, y, color=colors)

                for b, ano_i, mes_i in zip(bars, g["Ano"].tolist(), g["Mes"].tolist()):
                    b.set_picker(True)
                    self._pick_map[b] = {"type": "month_year", "value": (int(ano_i), int(mes_i))}

                self.ax1.set_title("Evolução do número de ocorrências (Mês/Ano) — Ano = Todos")
                self.ax1.set_xlabel("Mês/Ano")
                self.ax1.set_ylabel("Ocorrências")
                self.ax1.set_xticks(x)
                self.ax1.set_xticklabels(labels, rotation=45, ha="right")
                self.ax1.yaxis.set_major_locator(MaxNLocator(integer=True))

        # Ano específico => mantém Semanal/Mensal (com drilldown no Mensal)
        else:
            if self.granularidade.get() == "Semanal":
                mes_num_sel = self._mes_to_num(self.sel_mes.get())

                if mes_num_sel is not None:
                    dfp["SemanaMes"] = semana_no_mes(dfp[COL_DATA])
                    g = dfp.groupby("SemanaMes", as_index=False).size().rename(columns={"size": "Ocorrencias"})
                    full = pd.DataFrame({"SemanaMes": list(range(1, 7))})
                    g = full.merge(g, on="SemanaMes", how="left").fillna({"Ocorrencias": 0})

                    x = g["SemanaMes"].astype(int).tolist()
                    y = g["Ocorrencias"].astype(int).tolist()

                    colors = ["green" if v <= LIMIAR_SEMANAL else "red" for v in y]
                    bars = self.ax1.bar(x, y, color=colors)

                    for b, semm in zip(bars, x):
                        b.set_picker(True)
                        self._pick_map[b] = {"type": "week_month", "value": int(semm)}

                    ano_txt = self.sel_ano.get()
                    self.ax1.set_title(
                        f"Evolução (Semanal no mês) — {MESES_ABREV[int(mes_num_sel)]}/{ano_txt} | Limite: {LIMIAR_SEMANAL}"
                    )
                    self.ax1.set_xlabel("Semana do mês (1–6)")
                    self.ax1.set_ylabel("Ocorrências")
                    self.ax1.set_xlim(0.5, 6.5)
                    self.ax1.set_xticks(list(range(1, 7)))
                    self.ax1.yaxis.set_major_locator(MaxNLocator(integer=True))
                else:
                    dfp["Semana"] = semana_1a52(dfp[COL_DATA])
                    g = dfp.groupby("Semana", as_index=False).size().rename(columns={"size": "Ocorrencias"})
                    full = pd.DataFrame({"Semana": list(range(1, 53))})
                    g = full.merge(g, on="Semana", how="left").fillna({"Ocorrencias": 0})

                    x = g["Semana"].astype(int).tolist()
                    y = g["Ocorrencias"].astype(int).tolist()

                    colors = ["green" if v <= LIMIAR_SEMANAL else "red" for v in y]
                    bars = self.ax1.bar(x, y, color=colors)

                    for b, semana in zip(bars, x):
                        b.set_picker(True)
                        self._pick_map[b] = {"type": "week", "value": int(semana)}

                    self.ax1.set_title(f"Evolução do número de ocorrências (Semana 1–52) | Limite: {LIMIAR_SEMANAL}")
                    self.ax1.set_xlabel("Semana do ano")
                    self.ax1.set_ylabel("Ocorrências")
                    self.ax1.set_xlim(0.5, 52.5)
                    self.ax1.set_xticks(list(range(1, 53, 2)))
                    self.ax1.yaxis.set_major_locator(MaxNLocator(integer=True))

            else:
                # Mensal (Ano específico) => clique no mês abre semanas do mês
                ano = int(self.sel_ano.get())
                dfp["Mes"] = dfp[COL_DATA].dt.month.astype(int)
                g = dfp.groupby("Mes", as_index=False).size().rename(columns={"size": "Ocorrencias"})
                full = pd.DataFrame({"Mes": list(range(1, 13))})
                g = full.merge(g, on="Mes", how="left").fillna({"Ocorrencias": 0})

                labels = [mes_abrev(m) for m in g["Mes"].astype(int).tolist()]
                y = g["Ocorrencias"].astype(int).tolist()
                x = list(range(len(labels)))

                colors = ["green" if v <= LIMIAR_MENSAL else "red" for v in y]
                bars = self.ax1.bar(x, y, color=colors)

                for b, m in zip(bars, g["Mes"].astype(int).tolist()):
                    b.set_picker(True)
                    self._pick_map[b] = {"type": "month_drill", "value": int(m)}

                self.ax1.set_title(f"Evolução do número de ocorrências (Mensal — {ano}) | Limite: {LIMIAR_MENSAL}")
                self.ax1.set_xlabel("Mês (clique para ver semanas)")
                self.ax1.set_ylabel("Ocorrências")
                self.ax1.set_xticks(x)
                self.ax1.set_xticklabels(labels)
                self.ax1.yaxis.set_major_locator(MaxNLocator(integer=True))

        self.ax1.grid(True, axis="y", linestyle="--", linewidth=0.5, alpha=0.7)

        # -------- GRÁFICO MOTIVOS (ax2) --------
        self.ax2.clear()
        motivos = contar_top(self.df_filtrado, COL_MOTIVO, top_n=10, vazio="SEM MOTIVO")
        if len(motivos) > 0:
            bars2 = self.ax2.barh(motivos.index[::-1], motivos.values[::-1])
            for b, motivo in zip(bars2, motivos.index[::-1].tolist()):
                b.set_picker(True)
                self._pick_map[b] = {"type": "motivo", "value": motivo}

        self.ax2.set_title("Motivo da reclamação (Top 10 + Outros)")
        self.ax2.set_xlabel("Ocorrências")
        self.ax2.grid(True, axis="x", linestyle="--", linewidth=0.5, alpha=0.7)
        self.ax2.xaxis.set_major_locator(MaxNLocator(integer=True))

        self.fig.tight_layout()
        self.canvas_plot.draw()

        # -------- TABELA --------
        for item in self.tree.get_children():
            self.tree.delete(item)

        df_list = self.df_filtrado.sort_values(COL_DATA, ascending=False)
        for _, row in df_list.iterrows():
            status = row.get(COL_STATUS, "")
            tag = "normal"
            if is_reprovada(status):
                tag = "reprovada"
            elif is_associada(status):
                tag = "associada"

            self.tree.insert(
                "", tk.END,
                values=(
                    row.get(COL_CODIGO, ""),
                    row.get(COL_TITULO, ""),
                    status,
                    br_date_str(row.get(COL_DATA)),
                    row.get(COL_TURNO, ""),
                ),
                tags=(tag,)
            )

        total = len(self.df_filtrado)
        p_ini = br_date_str(self.df_filtrado[COL_DATA].min()) if total else "-"
        p_fim = br_date_str(self.df_filtrado[COL_DATA].max()) if total else "-"
        self.lbl_info.config(text=f"Registros no filtro: {total} | Período: {p_ini} a {p_fim}")

    # =========================================================
    # Clique em barras
    # =========================================================
    def _on_pick(self, event):
        artist = event.artist
        if artist not in self._pick_map:
            return

        info = self._pick_map[artist]
        ptype = info["type"]
        pval = info["value"]
        dfp = self.df_filtrado.copy()

        if ptype == "week":
            dfp["Semana"] = semana_1a52(dfp[COL_DATA])
            dfp = dfp[dfp["Semana"] == int(pval)]
            titulo = f"Ocorrências — Semana {pval}"
            self._abrir_popup_agrupado_por_responsavel(titulo, dfp)

        elif ptype == "week_month":
            dfp["SemanaMes"] = semana_no_mes(dfp[COL_DATA])
            dfp = dfp[dfp["SemanaMes"] == int(pval)]
            titulo = f"Ocorrências — Semana {pval} do mês"
            self._abrir_popup_agrupado_por_responsavel(titulo, dfp)

        elif ptype == "month_drill":
            ano = int(self.sel_ano.get())
            mes = int(pval)
            df_mes = dfp[(dfp[COL_DATA].dt.year == ano) & (dfp[COL_DATA].dt.month == mes)].copy()
            titulo = f"Ocorrências — {MESES_ABREV[mes]}/{ano} (Semanas do mês)"
            self._abrir_popup_semanas_do_mes(titulo, df_mes, ano, mes)

        elif ptype == "month_year":
            ano, mes = pval
            df_mes = dfp[(dfp[COL_DATA].dt.year == int(ano)) & (dfp[COL_DATA].dt.month == int(mes))].copy()
            titulo = f"Ocorrências — {MESES_ABREV[int(mes)]}/{int(ano)}"
            self._abrir_popup_agrupado_por_responsavel(titulo, df_mes)

        elif ptype == "motivo":
            motivo = str(pval)
            dfp = dfp[dfp[COL_MOTIVO].fillna("").replace("", "SEM MOTIVO") == motivo]
            titulo = f"Ocorrências — Motivo: {motivo}"
            self._abrir_popup_agrupado_por_responsavel(titulo, dfp)

    # =========================================================
    # Popup: semanas do mês + clique na semana abre lista detalhada
    # =========================================================
    def _abrir_popup_semanas_do_mes(self, titulo, df_mes, ano, mes):
        win = tk.Toplevel(self)
        win.title(titulo)
        win.geometry("1100x640")
        win.transient(self.winfo_toplevel())
        win.grab_set()

        topo = ttk.Frame(win, padding=10)
        topo.pack(fill="x")
        ttk.Label(topo, text=titulo, font=("Segoe UI", 10, "bold")).pack(side="left")
        ttk.Button(topo, text="Fechar", command=win.destroy).pack(side="right")

        # dados por semana (1..6)
        if df_mes is None or df_mes.empty:
            g = pd.DataFrame({"SemanaMes": list(range(1, 7)), "Ocorrencias": [0]*6})
        else:
            d = df_mes.copy()
            d["SemanaMes"] = semana_no_mes(d[COL_DATA])
            g = d.groupby("SemanaMes", as_index=False).size().rename(columns={"size": "Ocorrencias"})
            full = pd.DataFrame({"SemanaMes": list(range(1, 7))})
            g = full.merge(g, on="SemanaMes", how="left").fillna({"Ocorrencias": 0})

        x = g["SemanaMes"].astype(int).tolist()
        y = g["Ocorrencias"].astype(int).tolist()
        colors = ["green" if v <= LIMIAR_SEMANAL else "red" for v in y]

        pick_week = {}

        fig = Figure(figsize=(10.2, 4.0), dpi=110)
        ax = fig.add_subplot(111)
        bars = ax.bar(x, y, color=colors)
        for b, sem in zip(bars, x):
            b.set_picker(True)
            pick_week[b] = int(sem)

        ax.set_title(f"Ocorrências por semana do mês — {MESES_ABREV[int(mes)]}/{int(ano)} (clique na semana)")
        ax.set_xlabel("Semana do mês (1–6)")
        ax.set_ylabel("Ocorrências")
        ax.set_xlim(0.5, 6.5)
        ax.set_xticks(list(range(1, 7)))
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.grid(True, axis="y", linestyle="--", linewidth=0.5, alpha=0.7)

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        canvas.draw()

        # clique na semana -> abre lista detalhada daquele recorte (semana do mês)
        def on_pick_week(evt):
            art = evt.artist
            if art not in pick_week:
                return
            sem = pick_week[art]
            dff = df_mes.copy()
            dff["SemanaMes"] = semana_no_mes(dff[COL_DATA])
            dff = dff[dff["SemanaMes"] == int(sem)]
            self._abrir_popup_agrupado_por_responsavel(
                f"Ocorrências — {MESES_ABREV[int(mes)]}/{int(ano)} | Semana {sem} do mês",
                dff
            )

        canvas.mpl_connect("pick_event", on_pick_week)

        btns = ttk.Frame(win, padding=(10, 0, 10, 10))
        btns.pack(fill="x")

        ttk.Button(
            btns,
            text="Abrir lista detalhada do mês (por responsável)",
            command=lambda: self._abrir_popup_agrupado_por_responsavel(
                f"Lista — {MESES_ABREV[int(mes)]}/{int(ano)}", df_mes
            )
        ).pack(side="left")

    # =========================================================
    # Popup: lista agrupada por responsável (já existia)
    # =========================================================
    def _abrir_popup_agrupado_por_responsavel(self, titulo, df):
        win = tk.Toplevel(self)
        win.title(titulo)
        win.geometry("1220x680")
        win.transient(self.winfo_toplevel())
        win.grab_set()

        topo = ttk.Frame(win, padding=10)
        topo.pack(fill="x")
        ttk.Label(topo, text=titulo, font=("Segoe UI", 10, "bold")).pack(side="left")

        btns = ttk.Frame(topo)
        btns.pack(side="right")
        btn_export = ttk.Button(btns, text="Gerar Excel (este recorte)")
        btn_export.pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Fechar", command=win.destroy).pack(side="left")

        desc = ttk.Label(
            win,
            text="Agrupado por Responsável da análise de causa. Use o filtro abaixo para ver Todas ou Somente Atrasadas.",
            padding=(10, 0, 10, 8)
        )
        desc.pack(fill="x")

        filtro_frame = ttk.Frame(win, padding=(10, 0, 10, 10))
        filtro_frame.pack(fill="x")

        ttk.Label(filtro_frame, text="Filtro (Situação):").pack(side="left")
        var_popup_situacao = tk.StringVar(value="Todas")
        cb_popup = ttk.Combobox(
            filtro_frame,
            textvariable=var_popup_situacao,
            state="readonly",
            values=["Todas", "Somente ATRASADAS"],
            width=18
        )
        cb_popup.pack(side="left", padx=8)

        frame = ttk.Frame(win, padding=(10, 0, 10, 10))
        frame.pack(fill="both", expand=True)

        cols = ("Número", "Título", "Motivo", "Etapa", "Data de início", "Situação")
        tree = ttk.Treeview(frame, columns=cols, show="tree headings")

        tree.heading("#0", text="Responsável da análise de causa")
        tree.column("#0", width=240, anchor="w", stretch=True)
        for c in cols:
            tree.heading(c, text=c)

        tree.column("Número", width=80, anchor="w", stretch=False)
        tree.column("Título", width=330, anchor="w", stretch=True)
        tree.column("Motivo", width=180, anchor="w", stretch=True)
        tree.column("Etapa", width=110, anchor="w", stretch=False)
        tree.column("Data de início", width=110, anchor="w", stretch=False)
        tree.column("Situação", width=110, anchor="center", stretch=False)

        sb_y = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        sb_x = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)

        tree.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")

        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        tree.tag_configure("atrasada", foreground="red")
        tree.tag_configure("noprazo", foreground="green")
        tree.tag_configure("reprovada", foreground="blue")
        tree.tag_configure("associada", foreground="#A9A9A9")

        df_local = df.copy()
        if COL_RESP_ANALISE in df_local.columns:
            df_local["_RespAnalise"] = df_local[COL_RESP_ANALISE].fillna("").replace("", "SEM RESPONSÁVEL")
        else:
            df_local["_RespAnalise"] = "SEM RESPONSÁVEL"

        if COL_SITUACAO in df_local.columns:
            df_local["_Situacao"] = df_local[COL_SITUACAO].fillna("").apply(normalizar_situacao)
        else:
            df_local["_Situacao"] = ""

        dff_atual = {"df": df_local.copy()}

        def tags_linha(status_txt: str, situacao_txt: str):
            tags = []
            if situacao_txt == "ATRASADA":
                tags.append("atrasada")
            elif situacao_txt == "NO PRAZO":
                tags.append("noprazo")
            if is_reprovada(status_txt):
                tags.append("reprovada")
            elif is_associada(status_txt):
                tags.append("associada")
            return tuple(tags)

        lbl_total = ttk.Label(win, text="", padding=10)
        lbl_total.pack(fill="x")

        def preencher_tree():
            for item in tree.get_children():
                tree.delete(item)

            dff = df_local.copy()
            if var_popup_situacao.get() == "Somente ATRASADAS":
                dff = dff[dff["_Situacao"] == "ATRASADA"]

            dff = dff.sort_values(["_RespAnalise", COL_DATA], ascending=[True, False])
            dff_atual["df"] = dff

            grupos = {}
            for resp in dff["_RespAnalise"].unique().tolist():
                parent = tree.insert("", tk.END, text=resp, open=True)
                grupos[resp] = parent

            for _, r in dff.iterrows():
                parent = grupos.get(r["_RespAnalise"])
                sit = r.get("_Situacao", "")
                status = r.get(COL_STATUS, "")
                tree.insert(
                    parent,
                    tk.END,
                    text="",
                    values=(
                        r.get(COL_CODIGO, ""),
                        r.get(COL_TITULO, ""),
                        r.get(COL_MOTIVO, ""),
                        status,
                        br_date_str(r.get(COL_DATA)),
                        sit
                    ),
                    tags=tags_linha(status, sit)
                )

            lbl_total.config(text=f"Total de ocorrências (neste recorte): {len(dff)}")

        def exportar_popup():
            try:
                dff = dff_atual["df"].copy()
                if dff.empty:
                    messagebox.showinfo("Exportar", "Não há registros neste recorte para exportar.")
                    return

                cols_out = [COL_CODIGO, COL_TITULO, COL_STATUS, COL_DATA, COL_TURNO, COL_MOTIVO]
                for c in [COL_RESP_OCORRENCIA, COL_RESP_ANALISE, COL_SITUACAO]:
                    if c in dff.columns and c not in cols_out:
                        cols_out.append(c)

                out = dff[cols_out].copy() if all(c in dff.columns for c in cols_out) else dff.copy()
                if COL_DATA in out.columns:
                    out = out.sort_values(COL_DATA, ascending=False)
                    out[COL_DATA] = out[COL_DATA].apply(br_date_str)

                nome_padrao = f"Recorte_{pd.Timestamp.today().strftime('%d-%m-%Y')}.xlsx"
                caminho = filedialog.asksaveasfilename(
                    title="Salvar recorte do popup",
                    defaultextension=".xlsx",
                    initialfile=nome_padrao,
                    filetypes=[("Excel", "*.xlsx")]
                )
                if not caminho:
                    return

                with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
                    out.to_excel(writer, sheet_name="Recorte", index=False)

                messagebox.showinfo("Exportar", f"Arquivo gerado com sucesso:\n{caminho}")
            except Exception as e:
                messagebox.showerror("Erro ao exportar", str(e))

        btn_export.configure(command=exportar_popup)
        cb_popup.bind("<<ComboboxSelected>>", lambda e: preencher_tree())
        preencher_tree()


# =========================================================
# Abas genéricas (mantidas como no seu código)
# =========================================================
def detectar_coluna_data(df: pd.DataFrame):
    candidatos = [
        "Data", "Data de emissão", "Data Emissão", "Data de Emissão",
        "Data Ocorrência", "Data da ocorrência", "Data da Ocorrência",
        "Emissão", "Emissao", "DATA"
    ]
    cols = list(df.columns)
    for c in candidatos:
        if c in cols:
            return c
    for c in cols:
        if "data" in str(c).lower():
            return c
    return None


class AbaGenericaFrame(ttk.Frame):
    def __init__(self, master, nome_aba: str, default_excel_name: str):
        super().__init__(master)
        self.nome_aba = nome_aba
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.caminho_excel = tk.StringVar(value=os.path.join(base_dir, default_excel_name))
        self.sheet = DEFAULT_SHEET

        self.df_base = pd.DataFrame()
        self.df_filtrado = pd.DataFrame()

        self.sel_ano = tk.StringVar(value="(Todos)")
        self.sel_mes = tk.StringVar(value="(Todos)")
        self.granularidade = tk.StringVar(value="Semanal")

        self.filters = {}
        self.col_data = None

        self._montar_ui()
        self._atualizar()

    def _montar_ui(self):
        topo = ttk.Frame(self, padding=10)
        topo.pack(fill="x")

        ttk.Label(topo, text="Arquivo Excel:").grid(row=0, column=0, sticky="w")
        ttk.Entry(topo, textvariable=self.caminho_excel, width=80).grid(row=0, column=1, padx=6)
        ttk.Button(topo, text="Procurar...", command=self._procurar_excel).grid(row=0, column=2, padx=(0, 10))
        ttk.Button(topo, text="Atualizar", command=self._atualizar).grid(row=0, column=3)
        topo.grid_columnconfigure(1, weight=1)

        cfg = ttk.Frame(self, padding=(10, 0, 10, 10))
        cfg.pack(fill="x")

        ttk.Label(cfg, text="Ano:").grid(row=0, column=0, sticky="e")
        self.cb_ano = ttk.Combobox(cfg, textvariable=self.sel_ano, state="readonly", width=10)
        self.cb_ano.grid(row=0, column=1, padx=6, sticky="w")

        ttk.Label(cfg, text="Mês:").grid(row=0, column=2, padx=(10, 0), sticky="e")
        self.cb_mes = ttk.Combobox(cfg, textvariable=self.sel_mes, state="readonly", width=10)
        self.cb_mes.grid(row=0, column=3, padx=6, sticky="w")

        ttk.Label(cfg, text="Evolução:").grid(row=0, column=4, padx=(10, 0), sticky="e")
        self.cb_gran = ttk.Combobox(cfg, textvariable=self.granularidade, state="readonly",
                                    values=["Semanal", "Mensal"], width=10)
        self.cb_gran.grid(row=0, column=5, padx=6, sticky="w")

        ttk.Button(cfg, text="Aplicar", command=self._render).grid(row=0, column=6, padx=(14, 0))

        self.paned = ttk.Panedwindow(self, orient="horizontal")
        self.paned.pack(fill="both", expand=True, padx=10, pady=10)

        self.filtros_frame = ttk.LabelFrame(self.paned, text="Filtros por marcar", padding=8)
        self.painel_frame = ttk.Frame(self.paned)

        self.paned.add(self.filtros_frame, weight=1)
        self.paned.add(self.painel_frame, weight=4)

        _, inner = criar_sidebar_scroll(self.filtros_frame, width=320)
        self._filters_inner = inner

        self.fig = Figure(figsize=(11.4, 6.2), dpi=110)
        gs = self.fig.add_gridspec(nrows=2, ncols=1, height_ratios=[3, 2])
        self.ax1 = self.fig.add_subplot(gs[0, 0])
        self.ax2 = self.fig.add_subplot(gs[1, 0])

        self.canvas_plot = FigureCanvasTkAgg(self.fig, master=self.painel_frame)
        self.canvas_plot.get_tk_widget().pack(fill="x", expand=False, pady=(0, 10))

        tabela_frame = ttk.LabelFrame(self.painel_frame, text="Registros (filtro atual)", padding=8)
        tabela_frame.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(tabela_frame, columns=(), show="headings", height=18)
        self.tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(tabela_frame, orient="vertical", command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=sb.set)

        self.lbl_info = ttk.Label(self, text="", padding=10)
        self.lbl_info.pack(fill="x")

    def _procurar_excel(self):
        arquivo = filedialog.askopenfilename(
            title=f"Selecione o Excel — {self.nome_aba}",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls")]
        )
        if arquivo:
            self.caminho_excel.set(arquivo)
            self._atualizar()

    def _atualizar(self):
        try:
            df = pd.read_excel(self.caminho_excel.get(), sheet_name=self.sheet).copy()
            for c in df.columns:
                if df[c].dtype == "object":
                    df[c] = df[c].astype(str).str.strip().replace("nan", "")

            col_data = detectar_coluna_data(df)
            if not col_data:
                raise ValueError(f"Não encontrei coluna de data na planilha de '{self.nome_aba}'.")
            self.col_data = col_data

            df[col_data] = pd.to_datetime(df[col_data], errors="coerce", dayfirst=True)
            df = df.dropna(subset=[col_data])

            self.df_base = df
            self._montar_filtros_laterais()
            self._popular_filtros_rapidos()
            self._render()
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def _montar_filtros_laterais(self):
        for w in self._filters_inner.winfo_children():
            w.destroy()
        self.filters.clear()

        df = self.df_base
        cd = self.col_data

        cols_candidatas = []
        for c in df.columns:
            if c == cd:
                continue
            if str(c).strip().lower() in ("unidade", "local"):
                continue
            if df[c].dtype == "object":
                nun = df[c].nunique(dropna=True)
                if 1 < nun <= 60:
                    cols_candidatas.append(c)

        cols_candidatas = cols_candidatas[:10]

        for col in cols_candidatas:
            f = ChecklistFilter(self._filters_inner, col, on_change=self._render, height=90)
            f.pack(fill="x", pady=4)
            self.filters[col] = f
            vals = sorted(df[col].dropna().astype(str).replace("nan", "").unique().tolist())
            vals = [v for v in vals if v != ""]
            f.set_values_preserve(vals)

    def _popular_filtros_rapidos(self):
        df = self.df_base
        cd = self.col_data

        anos = sorted(df[cd].dt.year.dropna().unique().tolist())
        self.cb_ano["values"] = ["(Todos)"] + [str(a) for a in anos]
        if self.sel_ano.get() not in self.cb_ano["values"]:
            self.sel_ano.set("(Todos)")

        self.cb_mes["values"] = ["(Todos)"] + [MESES_ABREV[m] for m in range(1, 13)]
        if self.sel_mes.get() not in self.cb_mes["values"]:
            self.sel_mes.set("(Todos)")

        if self.sel_ano.get() == "(Todos)":
            self.granularidade.set("Semanal")
            self.cb_gran["values"] = ["Semanal"]
        else:
            self.cb_gran["values"] = ["Semanal", "Mensal"]

    def _mes_to_num(self, mes_str):
        if mes_str == "(Todos)":
            return None
        return INV_MESES_ABREV.get(mes_str)

    def _aplicar_filtros(self):
        df_f = self.df_base.copy()
        cd = self.col_data

        if self.sel_ano.get() == "(Todos)":
            self.granularidade.set("Semanal")
            self.cb_gran["values"] = ["Semanal"]
        else:
            self.cb_gran["values"] = ["Semanal", "Mensal"]

        ano = self.sel_ano.get()
        mes_num = self._mes_to_num(self.sel_mes.get())

        if ano != "(Todos)":
            df_f = df_f[df_f[cd].dt.year == int(ano)]
        if mes_num is not None:
            df_f = df_f[df_f[cd].dt.month == int(mes_num)]

        for col, widget in self.filters.items():
            if col not in df_f.columns:
                continue
            selecionados = widget.get_selected()
            if not selecionados:
                return df_f.iloc[0:0]
            df_f = df_f[df_f[col].astype(str).isin(selecionados)]

        return df_f

    def _render(self):
        if self.df_base is None or self.df_base.empty:
            return

        self.df_filtrado = self._aplicar_filtros()
        cd = self.col_data

        self.ax1.clear()
        dfp = self.df_filtrado.copy()

        if self.granularidade.get() == "Semanal":
            mes_num_sel = self._mes_to_num(self.sel_mes.get())
            if mes_num_sel is not None:
                dfp["SemanaMes"] = semana_no_mes(dfp[cd])
                g = dfp.groupby("SemanaMes", as_index=False).size().rename(columns={"size": "Ocorrencias"})
                full = pd.DataFrame({"SemanaMes": list(range(1, 7))})
                g = full.merge(g, on="SemanaMes", how="left").fillna({"Ocorrencias": 0})

                x = g["SemanaMes"].astype(int).tolist()
                y = g["Ocorrencias"].astype(int).tolist()
                colors = ["green" if v <= LIMIAR_SEMANAL else "red" for v in y]
                self.ax1.bar(x, y, color=colors)

                ano_txt = self.sel_ano.get() if self.sel_ano.get() != "(Todos)" else "Todos"
                self.ax1.set_title(f"{self.nome_aba} — Semanal no mês ({MESES_ABREV[int(mes_num_sel)]}/{ano_txt}) | Limite: {LIMIAR_SEMANAL}")
                self.ax1.set_xlabel("Semana do mês (1–6)")
                self.ax1.set_ylabel("Ocorrências")
                self.ax1.set_xlim(0.5, 6.5)
                self.ax1.set_xticks(list(range(1, 7)))
                self.ax1.yaxis.set_major_locator(MaxNLocator(integer=True))
            else:
                dfp["Semana"] = semana_1a52(dfp[cd])
                g = dfp.groupby("Semana", as_index=False).size().rename(columns={"size": "Ocorrencias"})
                full = pd.DataFrame({"Semana": list(range(1, 53))})
                g = full.merge(g, on="Semana", how="left").fillna({"Ocorrencias": 0})

                x = g["Semana"].astype(int).tolist()
                y = g["Ocorrencias"].astype(int).tolist()
                colors = ["green" if v <= LIMIAR_SEMANAL else "red" for v in y]
                self.ax1.bar(x, y, color=colors)

                self.ax1.set_title(f"{self.nome_aba} — Ocorrências por Semana (1–52) | Limite: {LIMIAR_SEMANAL}")
                self.ax1.set_xlabel("Semana")
                self.ax1.set_ylabel("Ocorrências")
                self.ax1.set_xlim(0.5, 52.5)
                self.ax1.set_xticks(list(range(1, 53, 2)))
                self.ax1.yaxis.set_major_locator(MaxNLocator(integer=True))
        else:
            ano = int(self.sel_ano.get())
            dfp["Mes"] = dfp[cd].dt.month.astype(int)
            g = dfp.groupby("Mes", as_index=False).size().rename(columns={"size": "Ocorrencias"})
            full = pd.DataFrame({"Mes": list(range(1, 13))})
            g = full.merge(g, on="Mes", how="left").fillna({"Ocorrencias": 0})

            labels = [mes_abrev(m) for m in g["Mes"].astype(int).tolist()]
            y = g["Ocorrencias"].astype(int).tolist()
            x = list(range(len(labels)))
            colors = ["green" if v <= LIMIAR_MENSAL else "red" for v in y]
            self.ax1.bar(x, y, color=colors)

            self.ax1.set_title(f"{self.nome_aba} — Ocorrências Mensais ({ano}) | Limite: {LIMIAR_MENSAL}")
            self.ax1.set_xlabel("Mês")
            self.ax1.set_ylabel("Ocorrências")
            self.ax1.set_xticks(x)
            self.ax1.set_xticklabels(labels)
            self.ax1.yaxis.set_major_locator(MaxNLocator(integer=True))

        self.ax1.grid(True, axis="y", linestyle="--", linewidth=0.5, alpha=0.7)

        self.ax2.clear()
        col_cat = None
        for c in self.df_base.columns:
            if c == cd:
                continue
            if str(c).strip().lower() in ("unidade", "local"):
                continue
            if self.df_base[c].dtype == "object" and self.df_base[c].nunique(dropna=True) > 1:
                col_cat = c
                break

        if col_cat:
            top = contar_top(self.df_filtrado, col_cat, top_n=10, vazio="SEM VALOR")
            if len(top) > 0:
                self.ax2.barh(top.index[::-1], top.values[::-1])
            self.ax2.set_title(f"Top 10 — {col_cat}")
            self.ax2.set_xlabel("Ocorrências")
            self.ax2.grid(True, axis="x", linestyle="--", linewidth=0.5, alpha=0.7)
            self.ax2.xaxis.set_major_locator(MaxNLocator(integer=True))
        else:
            self.ax2.set_title("Top 10 (sem coluna categórica adequada)")
            self.ax2.text(0.5, 0.5, "Sem dados categóricos para o Top 10", ha="center", va="center")

        self.fig.tight_layout()
        self.canvas_plot.draw()

        cols = list(self.df_base.columns)[:12]
        self.tree["columns"] = tuple(cols)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=160, anchor="w")

        for item in self.tree.get_children():
            self.tree.delete(item)

        df_list = self.df_filtrado.sort_values(cd, ascending=False)
        if len(df_list) > 5000:
            df_list = df_list.head(5000)

        for _, row in df_list.iterrows():
            vals = []
            for c in cols:
                v = row.get(c, "")
                if pd.isna(v):
                    v = ""
                if isinstance(v, pd.Timestamp):
                    v = v.strftime(DATE_FMT_BR)
                vals.append(str(v))
            self.tree.insert("", tk.END, values=tuple(vals))

        total = len(self.df_filtrado)
        p_ini = br_date_str(self.df_filtrado[cd].min()) if total else "-"
        p_fim = br_date_str(self.df_filtrado[cd].max()) if total else "-"
        self.lbl_info.config(text=f"{self.nome_aba} | Registros no filtro: {total} | Período: {p_ini} a {p_fim}")


# =========================
# App com abas
# =========================
class IndicadoresQualidadeRS(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("INDICADORES QUALIDADE RS")
        self.geometry("1500x920")
        self.minsize(1200, 720)

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True)

        nb.add(ReclamacoesClienteFrame(nb), text="Reclamações de Cliente")
        nb.add(AbaGenericaFrame(nb, "Desvios", "Coleta Desvios.xlsx"), text="Desvios")
        nb.add(AbaGenericaFrame(nb, "Retrabalho", "Coleta Retrabalho.xlsx"), text="Retrabalho")
        nb.add(AbaGenericaFrame(nb, "Performance", "Coleta Performance.xlsx"), text="Performance")
        nb.add(AbaGenericaFrame(nb, "Pessoas", "Coleta Pessoas.xlsx"), text="Pessoas")


if __name__ == "__main__":
    IndicadoresQualidadeRS().mainloop()