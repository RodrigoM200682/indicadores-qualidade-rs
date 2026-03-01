# app.py — INDICADORES QUALIDADE RS (WEB) — V18.02.26
# Atualização desta versão:
# - Painel interativo (Ocorrências) agora aparece APENAS UMA VEZ e fica LADO A LADO com Motivos (Top 12)
# - Mantém: drill (Ano->Mês->Semana), botão Voltar, Reset drill, Tabela por barra clicada, Pizza e Atrasadas

import io
import os
import re
import json
import pandas as pd
import streamlit as st
import plotly.express as px

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.chart import BarChart, PieChart
from openpyxl.chart.reference import Reference
from openpyxl.chart.label import DataLabelList

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader


# =========================================================
# APP
# =========================================================
APP_VERSION = "V22.02.26_T"
APP_NAME = f"INDICADORES QUALIDADE RS — {APP_VERSION}"
DEFAULT_SHEET = "Sheet1"
APP_PASSWORD = "QualidadeRS"

# Colunas esperadas (base Reclamações)
COL_CODIGO = "Código"
COL_TITULO = "Título"
COL_STATUS = "Status"
COL_DATA = "Data de emissão"
COL_MOTIVO = "Motivo Reclamação"
COL_TURNO = "Turno/Horário"
COL_RESP_OCORRENCIA = "Responsável"
COL_RESP_ANALISE = "Responsável da análise de causa"
COL_CATEGORIA = "Categoria"
COL_SITUACAO = "Situação"  # ATRASADA / NO PRAZO

# ✅ Filtros por marcar (com Categoria incluída)
FILTROS_COLS = [
    "Status",
    "Categoria",
    "Cliente",
    "Motivo Reclamação",
    "Responsável",
    "Responsável da análise de causa",
    "Turno/Horário",
    "Embalagem",
    COL_SITUACAO,
]

DATE_FMT_BR = "%d/%m/%Y"
MESES_ABREV = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
}
INV_MESES_ABREV = {v: k for k, v in MESES_ABREV.items()}

# Regras de cores
LIMIAR_OCORRENCIAS = 8  # <=8 verde, >8 vermelho
GREEN = "#2E7D32"
RED = "#C62828"
BLUE = "#1f77b4"


# =========================================================
# Login (senha fixa)
# =========================================================
def require_login():
    if "auth_ok" not in st.session_state:
        st.session_state.auth_ok = False

    if st.session_state.auth_ok:
        return

    st.title(f"🔒 {APP_NAME}")
    st.write("Acesso interno. Informe a senha para continuar.")
    p = st.text_input("Senha", type="password")

    if st.button("Entrar"):
        if p == APP_PASSWORD:
            st.session_state.auth_ok = True
            st.rerun()
        else:
            st.error("Senha inválida.")

    st.stop()


# =========================================================
# Utilitários
# =========================================================
def br_date_str(dt) -> str:
    if pd.isna(dt):
        return ""
    try:
        return pd.to_datetime(dt).strftime(DATE_FMT_BR)
    except Exception:
        return ""


def normalizar_situacao(x: str) -> str:
    s = str(x).strip().upper()
    if s in ("ATRASADA", "ATRASADO"):
        return "ATRASADA"
    if s in ("NO PRAZO", "NOPRAZO", "NO_PRAZO"):
        return "NO PRAZO"
    if s in ("", "NAN", "NONE"):
        return ""
    return s


def _titulo_filtro(ano_sel: str, mes_sel: str, resp_occ_sel: str) -> str:
    return f"Ano {ano_sel} | Mês {mes_sel} | Resp ocorrência {resp_occ_sel}"


def semana_do_mes(dt_series: pd.Series) -> pd.Series:
    d = pd.to_datetime(dt_series, errors="coerce")
    return ((d.dt.day - 1) // 7 + 1).astype("Int64")


# =========================================================
# Excel helpers
# =========================================================
def _excel_theme():
    return {
        "title_fill": PatternFill("solid", fgColor="1F4E79"),
        "hdr_fill": PatternFill("solid", fgColor="2F5597"),
        "hdr_font": Font(bold=True, color="FFFFFF"),
        "thin": Side(style="thin", color="A6A6A6"),
        "kpi_fill": PatternFill("solid", fgColor="FFF2CC"),
    }


def _xl_col(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def _apply_border(ws, cell_range: str):
    th = _excel_theme()["thin"]
    border = Border(left=th, right=th, top=th, bottom=th)
    for row in ws[cell_range]:
        for c in row:
            c.border = border


def _merge_title(ws, cell_range: str, text: str):
    theme = _excel_theme()
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.fill = theme["title_fill"]
    c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _add_table(ws, start_row, start_col, df: pd.DataFrame, table_name: str, style="TableStyleMedium9"):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, v in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=v)

    end_row = start_row + len(df)
    end_col = start_col + df.shape[1] - 1
    ref = f"{_xl_col(start_col)}{start_row}:{_xl_col(end_col)}{end_row}"

    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tab)

    theme = _excel_theme()
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=c)
        cell.fill = theme["hdr_fill"]
        cell.font = theme["hdr_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.auto_filter.ref = ref
    _apply_border(ws, ref)
    return (start_row, start_col, end_row, end_col, ref)


def _hex_no_hash(hex_color: str) -> str:
    return hex_color.replace("#", "").upper()


def _style_xl_bar_chart(chart: BarChart, rotate_x_45: bool, solid_fill_hex: str | None):
    try:
        chart.y_axis.majorGridlines = None
        chart.y_axis.minorGridlines = None
        chart.x_axis.majorGridlines = None
        chart.x_axis.minorGridlines = None
    except Exception:
        pass
    try:
        chart.y_axis.tickLblPos = "none"
    except Exception:
        pass
    try:
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
    except Exception:
        pass
    if rotate_x_45:
        try:
            chart.x_axis.textRotation = 45
        except Exception:
            pass
    if solid_fill_hex:
        try:
            s = chart.series[0]
            s.graphicalProperties.solidFill = _hex_no_hash(solid_fill_hex)
            s.graphicalProperties.line.solidFill = _hex_no_hash(solid_fill_hex)
        except Exception:
            pass


def _style_xl_pie_chart(chart: PieChart):
    try:
        chart.dLbls = DataLabelList()
        chart.dLbls.showPercent = True
        chart.dLbls.showCatName = True
    except Exception:
        pass


def _add_bar_chart_from_sheet(
    data_ws, target_ws,
    title, cat_col, val_col, start_row, end_row, anchor_cell,
    rotate_x_45=False,
    height=7.2, width=12.5,
    solid_fill_hex=None
):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.legend = None
    chart.height = float(height)
    chart.width = float(width)

    data = Reference(data_ws, min_col=val_col, min_row=start_row, max_row=end_row)
    cats = Reference(data_ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    _style_xl_bar_chart(chart, rotate_x_45=rotate_x_45, solid_fill_hex=solid_fill_hex)
    target_ws.add_chart(chart, anchor_cell)


def _add_pie_chart_from_sheet(
    data_ws, target_ws,
    title, cat_col, val_col, start_row, end_row, anchor_cell,
    height=7.2, width=12.5
):
    chart = PieChart()
    chart.title = title
    chart.height = float(height)
    chart.width = float(width)

    data = Reference(data_ws, min_col=val_col, min_row=start_row, max_row=end_row)
    cats = Reference(data_ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    _style_xl_pie_chart(chart)
    target_ws.add_chart(chart, anchor_cell)


# =========================================================
# PDF do Dashboard (1 página) — Plotly -> PNG via kaleido
# =========================================================
def build_dashboard_pdf_bytes(app_name: str, filtro_txt: str, kpis: dict, figs_plotly: list) -> bytes:
    img_bytes_list = []
    for fig in figs_plotly:
        b = fig.to_image(format="png", scale=2)
        img_bytes_list.append(io.BytesIO(b))

    out = io.BytesIO()
    page_size = landscape(A4)
    c = canvas.Canvas(out, pagesize=page_size)
    W, H = page_size

    margin = 24
    header_h = 70
    footer_h = 18
    content_top = H - margin - header_h
    content_bottom = margin + footer_h
    content_h = content_top - content_bottom
    content_w = W - 2 * margin

    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, H - margin - 22, f"{app_name} — Dashboard")

    c.setFont("Helvetica", 9)
    c.drawString(margin, H - margin - 40, f"Filtro: {filtro_txt[:180]}")

    c.setFont("Helvetica-Bold", 10)
    kpi_line = (
        f"Total: {kpis.get('total', 0)}   |   "
        f"Em atraso: {kpis.get('atras', 0)}   |   "
        f"Período: {kpis.get('periodo', '-')}   |   "
        f"Versão: {APP_VERSION}"
    )
    c.drawString(margin, H - margin - 58, kpi_line)

    gap = 12
    cell_w = (content_w - gap) / 2
    cell_h = (content_h - gap) / 2

    positions = [(0, 0), (1, 0), (0, 1), (1, 1)]
    for i, (col, row) in enumerate(positions):
        if i >= len(img_bytes_list):
            break
        x = margin + col * (cell_w + gap)
        y = content_bottom + (1 - row) * (cell_h + gap)
        img = ImageReader(img_bytes_list[i])
        c.drawImage(img, x, y, width=cell_w, height=cell_h, preserveAspectRatio=True, anchor="c")

    c.setFont("Helvetica", 8)
    c.drawRightString(W - margin, margin / 2, f"Gerado em {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage()
    c.save()
    out.seek(0)
    return out.read()


# =========================================================
# Carregamento e filtros
# =========================================================
@st.cache_data(show_spinner=False)
def carregar_df(upload_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(upload_bytes), sheet_name=sheet_name)

    obrig = [COL_CODIGO, COL_TITULO, COL_STATUS, COL_DATA, COL_MOTIVO]
    for c in obrig:
        if c not in df.columns:
            raise ValueError(f"Não encontrei a coluna obrigatória '{c}' na planilha.")

    df = df.copy()
    df[COL_DATA] = pd.to_datetime(df[COL_DATA], errors="coerce", dayfirst=True)
    df = df.dropna(subset=[COL_DATA])

    for c in df.columns:
        if df[c].dtype == "object":
            df[c] = df[c].astype(str).str.strip().replace("nan", "")

    if COL_SITUACAO in df.columns:
        df[COL_SITUACAO] = df[COL_SITUACAO].apply(normalizar_situacao)

    return df


def aplicar_filtros(df: pd.DataFrame, ano_sel, mes_sel, resp_occ_sel, multi_filters: dict) -> pd.DataFrame:
    dff = df.copy()

    if ano_sel != "(Todos)":
        dff = dff[dff[COL_DATA].dt.year == int(ano_sel)]

    if mes_sel != "(Todos)":
        mes_num = INV_MESES_ABREV.get(mes_sel)
        if mes_num:
            dff = dff[dff[COL_DATA].dt.month == int(mes_num)]

    if resp_occ_sel != "(Todos)" and COL_RESP_OCORRENCIA in dff.columns:
        dff = dff[dff[COL_RESP_OCORRENCIA].astype(str) == resp_occ_sel]

    for col, selecionados in multi_filters.items():
        if col not in dff.columns:
            continue
        if not selecionados:
            return dff.iloc[0:0]
        dff = dff[dff[col].astype(str).isin(selecionados)]

    return dff


# =========================================================
# Drilldown + seleção da tabela
# =========================================================
def init_drill_state():
    if "drill_level" not in st.session_state:
        st.session_state.drill_level = "AUTO"  # AUTO / ANO / MES / SEMANA
    if "drill_year" not in st.session_state:
        st.session_state.drill_year = None
    if "drill_month" not in st.session_state:
        st.session_state.drill_month = None

    if "table_focus_level" not in st.session_state:
        st.session_state.table_focus_level = None  # "ANO"/"MES"/"SEMANA"
    if "table_focus_value" not in st.session_state:
        st.session_state.table_focus_value = None


def reset_drill():
    st.session_state.drill_level = "AUTO"
    st.session_state.drill_year = None
    st.session_state.drill_month = None
    st.session_state.table_focus_level = None
    st.session_state.table_focus_value = None


def clear_table_focus():
    st.session_state.table_focus_level = None
    st.session_state.table_focus_value = None


def resolve_initial_level(ano_sel: str, mes_sel: str):
    if ano_sel == "(Todos)":
        return "ANO"
    if mes_sel == "(Todos)":
        return "MES"
    return "SEMANA"


def apply_drill_filters(df_filtrado: pd.DataFrame, ano_sel: str, mes_sel: str) -> pd.DataFrame:
    dff = df_filtrado.copy()

    if ano_sel == "(Todos)" and st.session_state.drill_year is not None:
        dff = dff[dff[COL_DATA].dt.year == int(st.session_state.drill_year)]

    if mes_sel == "(Todos)" and st.session_state.drill_month is not None:
        dff = dff[dff[COL_DATA].dt.month == int(st.session_state.drill_month)]

    return dff


def apply_table_focus(df_context: pd.DataFrame) -> pd.DataFrame:
    lvl = st.session_state.table_focus_level
    val = st.session_state.table_focus_value
    if not lvl or val is None:
        return df_context

    dff = df_context.copy()

    if lvl == "ANO":
        try:
            y = int(val)
            dff = dff[dff[COL_DATA].dt.year == y]
        except Exception:
            return df_context

    elif lvl == "MES":
        try:
            m = INV_MESES_ABREV.get(str(val))
            if m:
                dff = dff[dff[COL_DATA].dt.month == int(m)]
        except Exception:
            return df_context

    elif lvl == "SEMANA":
        try:
            s = str(val).replace("ª", "").strip()
            w = int(s)
            dff = dff.copy()
            dff["_SEMANA_MES"] = semana_do_mes(dff[COL_DATA])
            dff = dff[dff["_SEMANA_MES"] == w].drop(columns=["_SEMANA_MES"], errors="ignore")
        except Exception:
            return df_context

    return dff


def occurrences_dataset(df_filtrado: pd.DataFrame, ano_sel: str, mes_sel: str):
    level = st.session_state.drill_level
    if level == "AUTO":
        level = resolve_initial_level(ano_sel, mes_sel)

    breadcrumb = []

    if level == "ANO":
        g = df_filtrado.groupby(df_filtrado[COL_DATA].dt.year).size().sort_index()
        df_plot = pd.DataFrame({"Ano": g.index.astype(int), "Ocorrências": g.values.astype(int)})
        breadcrumb.append("Visão: Ano")
        return df_plot, "ANO", " > ".join(breadcrumb)

    if ano_sel != "(Todos)":
        ano_alvo = int(ano_sel)
    else:
        ano_alvo = int(st.session_state.drill_year) if st.session_state.drill_year is not None else None

    if ano_alvo is None:
        g = df_filtrado.groupby(df_filtrado[COL_DATA].dt.year).size().sort_index()
        df_plot = pd.DataFrame({"Ano": g.index.astype(int), "Ocorrências": g.values.astype(int)})
        breadcrumb.append("Visão: Ano")
        return df_plot, "ANO", " > ".join(breadcrumb)

    breadcrumb.append(f"Ano {ano_alvo}")
    df_ano = df_filtrado[df_filtrado[COL_DATA].dt.year == ano_alvo].copy()

    if level == "MES":
        df_ano["MesNum"] = df_ano[COL_DATA].dt.month.astype(int)
        g = df_ano.groupby("MesNum").size().reindex(range(1, 13), fill_value=0)
        df_plot = pd.DataFrame({"Mês": [MESES_ABREV[m] for m in range(1, 13)], "Ocorrências": g.values.astype(int)})
        breadcrumb.append("Visão: Mês")
        return df_plot, "MES", " > ".join(breadcrumb)

    if mes_sel != "(Todos)":
        mes_alvo = int(INV_MESES_ABREV.get(mes_sel))
    else:
        mes_alvo = int(st.session_state.drill_month) if st.session_state.drill_month is not None else None

    if mes_alvo is None:
        df_ano["MesNum"] = df_ano[COL_DATA].dt.month.astype(int)
        g = df_ano.groupby("MesNum").size().reindex(range(1, 13), fill_value=0)
        df_plot = pd.DataFrame({"Mês": [MESES_ABREV[m] for m in range(1, 13)], "Ocorrências": g.values.astype(int)})
        breadcrumb.append("Visão: Mês")
        return df_plot, "MES", " > ".join(breadcrumb)

    breadcrumb.append(f"Mês {MESES_ABREV.get(mes_alvo, mes_alvo)}")
    breadcrumb.append("Visão: Semana do mês")

    df_mes = df_ano[df_ano[COL_DATA].dt.month == mes_alvo].copy()
    df_mes["Semana"] = semana_do_mes(df_mes[COL_DATA])
    g = df_mes.groupby("Semana").size().sort_index()

    idx = [1, 2, 3, 4, 5]
    g = g.reindex(idx, fill_value=0)
    df_plot = pd.DataFrame({"Semana": [f"{i}ª" for i in idx], "Ocorrências": g.values.astype(int)})

    return df_plot, "SEMANA", " > ".join(breadcrumb)


def get_clicked_x(plotly_event):
    if not plotly_event:
        return None
    try:
        sel = plotly_event.get("selection", {})
        pts = sel.get("points", [])
        if pts:
            return pts[0].get("x")
    except Exception:
        pass
    try:
        pts = plotly_event.get("points", [])
        if pts:
            return pts[0].get("x")
    except Exception:
        pass
    return None


def can_go_back(level_now: str, ano_sel: str, mes_sel: str) -> bool:
    if level_now == "SEMANA":
        return (mes_sel == "(Todos)") and (st.session_state.drill_month is not None)
    if level_now == "MES":
        return (ano_sel == "(Todos)") and (st.session_state.drill_year is not None)
    return False


def go_back_one_level(level_now: str, ano_sel: str, mes_sel: str):
    if level_now == "SEMANA" and mes_sel == "(Todos)" and st.session_state.drill_month is not None:
        st.session_state.drill_level = "MES"
        st.session_state.drill_month = None
        clear_table_focus()
        return True
    if level_now == "MES" and ano_sel == "(Todos)" and st.session_state.drill_year is not None:
        st.session_state.drill_level = "ANO"
        st.session_state.drill_year = None
        st.session_state.drill_month = None
        clear_table_focus()
        return True
    return False


# =========================================================
# Datasets (seguindo seleção)
# =========================================================
def calc_resp_analise(df_context: pd.DataFrame):
    resp = (
        df_context[COL_RESP_ANALISE].fillna("").replace("", "SEM RESPONSÁVEL")
        if COL_RESP_ANALISE in df_context.columns else pd.Series(["SEM RESPONSÁVEL"] * len(df_context))
    )
    df_resp = resp.value_counts().reset_index()
    df_resp.columns = ["Responsável (análise)", "Ocorrências"]
    if df_resp.empty:
        df_resp = pd.DataFrame({"Responsável (análise)": ["SEM DADOS"], "Ocorrências": [0]})
    return df_resp


def calc_motivos(df_context: pd.DataFrame, top_n=12):
    top_mot = (
        df_context[COL_MOTIVO].fillna("").replace("", "SEM MOTIVO").value_counts().head(top_n)
        if COL_MOTIVO in df_context.columns else pd.Series(dtype=int)
    )
    df_mot = top_mot.reset_index()
    df_mot.columns = ["Motivo", "Ocorrências"]
    if df_mot.empty:
        df_mot = pd.DataFrame({"Motivo": ["SEM DADOS"], "Ocorrências": [0]})
    return df_mot


def calc_atrasadas_por_filtro(df_filtro_base: pd.DataFrame):
    dfb = df_filtro_base.copy()

    resp = (
        dfb[COL_RESP_ANALISE].fillna("").replace("", "SEM RESPONSÁVEL")
        if COL_RESP_ANALISE in dfb.columns else pd.Series(["SEM RESPONSÁVEL"] * len(dfb))
    )
    sit = (
        dfb[COL_SITUACAO].fillna("").apply(normalizar_situacao)
        if COL_SITUACAO in dfb.columns else pd.Series([""] * len(dfb))
    )

    df_atras = (
        pd.DataFrame({"Responsável (análise)": resp, "Situação": sit})
        .query("Situação == 'ATRASADA'")["Responsável (análise)"]
        .value_counts()
        .reset_index()
    )
    df_atras.columns = ["Responsável (análise)", "Atrasadas (filtro)"]
    if df_atras.empty:
        df_atras = pd.DataFrame({"Responsável (análise)": ["SEM DADOS"], "Atrasadas (filtro)": [0]})
    return df_atras


# =========================================================
# Plotly styling (sem eixo Y)
# =========================================================
def _hide_yaxis(fig):
    fig.update_yaxes(showticklabels=False, title=None, showgrid=True)
    fig.update_layout(showlegend=False)
    return fig


def _common_bar_layout(fig, height=460):
    fig.update_layout(
        height=height,
        margin=dict(l=10, r=10, t=55, b=10),
    )
    return fig


def _apply_threshold_colors(fig, values, threshold: int):
    colors = [GREEN if int(v) <= threshold else RED for v in values]
    fig.update_traces(marker_color=colors)
    return fig


def fig_ocorrencias(df_plot: pd.DataFrame, level: str):
    if level == "ANO":
        fig = px.bar(df_plot, x="Ano", y="Ocorrências", title="Ocorrências (clique para detalhar)")
        fig.update_traces(text=df_plot["Ocorrências"], textposition="outside", cliponaxis=False)
        _apply_threshold_colors(fig, df_plot["Ocorrências"].tolist(), LIMIAR_OCORRENCIAS)
        _hide_yaxis(fig)
        _common_bar_layout(fig, height=460)
        return fig

    if level == "MES":
        fig = px.bar(df_plot, x="Mês", y="Ocorrências", title="Ocorrências por mês (clique para detalhar por semana)")
        fig.update_traces(text=df_plot["Ocorrências"], textposition="outside", cliponaxis=False)
        _apply_threshold_colors(fig, df_plot["Ocorrências"].tolist(), LIMIAR_OCORRENCIAS)
        _hide_yaxis(fig)
        _common_bar_layout(fig, height=460)
        return fig

    fig = px.bar(df_plot, x="Semana", y="Ocorrências", title="Ocorrências por semana do mês (clique para ver tabela)")
    fig.update_traces(text=df_plot["Ocorrências"], textposition="outside", cliponaxis=False)
    _apply_threshold_colors(fig, df_plot["Ocorrências"].tolist(), LIMIAR_OCORRENCIAS)
    _hide_yaxis(fig)
    _common_bar_layout(fig, height=460)
    return fig


def fig_motivos(df_mot: pd.DataFrame, titulo: str):
    fig = px.bar(df_mot, x="Motivo", y="Ocorrências", title=titulo)
    fig.update_traces(text=df_mot["Ocorrências"], textposition="outside", cliponaxis=False, marker_color=BLUE)
    fig.update_layout(xaxis_tickangle=-45)
    _hide_yaxis(fig)
    _common_bar_layout(fig, height=460)
    return fig


def fig_pizza_participacao(df_resp: pd.DataFrame, titulo: str):
    """Gráfico de participação por responsável (análise) em BARRAS (substitui pizza)."""
    ycol = df_resp["Ocorrências"] if "Ocorrências" in df_resp.columns else df_resp.iloc[:, 1]
    fig = px.bar(df_resp, x="Responsável (análise)", y=ycol.name if hasattr(ycol,'name') and ycol.name else "Ocorrências", title=titulo)
    # rótulos e layout
    try:
        fig.update_traces(text=df_resp[ycol.name if hasattr(ycol,'name') and ycol.name else "Ocorrências"], textposition="outside", cliponaxis=False)
    except Exception:
        pass
    fig.update_layout(xaxis_tickangle=-45)
    _hide_yaxis(fig)
    fig.update_layout(height=420, margin=dict(l=10, r=10, t=55, b=10))
    return fig


def fig_atrasadas_vermelho(df_atras: pd.DataFrame, titulo: str):
    ycol = df_atras.columns[1]
    fig = px.bar(df_atras, x="Responsável (análise)", y=ycol, title=titulo)
    fig.update_traces(text=df_atras[ycol], textposition="outside", cliponaxis=False, marker_color=RED)
    fig.update_layout(xaxis_tickangle=-45)
    _hide_yaxis(fig)
    fig.update_layout(height=420, margin=dict(l=10, r=10, t=55, b=10))
    return fig


# =========================================================
# Resumo Excel (DASHBOARD + DADOS + RECORTE) — com Pizza
# (mesmo da sua versão anterior; mantido para não quebrar export)
# =========================================================
def build_resumo_excel_bytes(df_filtrado_final: pd.DataFrame, df_filtro_base: pd.DataFrame, titulo_filtro: str) -> bytes:
    dff = df_filtrado_final.copy()
    theme = _excel_theme()

    total_rec = int(len(dff))
    p_ini = br_date_str(dff[COL_DATA].min()) if total_rec else "-"
    p_fim = br_date_str(dff[COL_DATA].max()) if total_rec else "-"

    if COL_SITUACAO in dff.columns and total_rec:
        sit_rec = dff[COL_SITUACAO].fillna("").apply(normalizar_situacao)
        total_atras_rec = int((sit_rec == "ATRASADA").sum())
    else:
        total_atras_rec = 0

    dff_mes = dff.copy()
    dff_mes["MesNum"] = dff_mes[COL_DATA].dt.month.astype(int)
    g_mes = dff_mes.groupby("MesNum").size().reindex(range(1, 13), fill_value=0)
    df_mes = pd.DataFrame({"Mês": [MESES_ABREV[m] for m in range(1, 13)], "Ocorrências": g_mes.values.astype(int)})

    df_resp = calc_resp_analise(dff)
    df_atras = calc_atrasadas_por_filtro(df_filtro_base)
    df_mot = calc_motivos(dff, top_n=12)

    wb = Workbook()
    ws = wb.active
    ws.title = "DASHBOARD"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    _set_col_widths(ws, {"A": 2, "B": 28, "C": 28, "D": 28, "E": 28, "F": 2})

    _merge_title(ws, "B2:E2", "RESUMO DE OCORRÊNCIAS — DASHBOARD (4 GRÁFICOS)")
    ws.row_dimensions[2].height = 26

    ws["B4"] = "Filtro:"; ws["B4"].font = Font(bold=True)
    ws["C4"] = titulo_filtro
    ws.merge_cells("C4:E4")
    ws["C4"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["B6"] = "Período (recorte):"; ws["B6"].font = Font(bold=True)
    ws["C6"] = f"{p_ini} a {p_fim}"
    ws["D6"] = "Versão:"; ws["D6"].font = Font(bold=True)
    ws["E6"] = APP_VERSION

    for row in ws["B8:E10"]:
        for c in row:
            c.fill = theme["kpi_fill"]
            c.alignment = Alignment(vertical="center", wrap_text=True)
    _apply_border(ws, "B8:E10")

    ws["B8"] = "TOTAIS (RECORTE FINAL)"; ws["B8"].font = Font(bold=True, size=12); ws.merge_cells("B8:E8")
    ws["B9"] = "Total de ocorrências"; ws["B9"].font = Font(bold=True)
    ws["C9"] = total_rec; ws["C9"].font = Font(bold=True, size=14)
    ws["D9"] = "Ocorrências em atraso (recorte)"; ws["D9"].font = Font(bold=True)
    ws["E9"] = total_atras_rec; ws["E9"].font = Font(bold=True, size=14)
    ws["B10"] = "Obs.: tabelas base ficam na aba DADOS."; ws.merge_cells("B10:E10")

    wsd = wb.create_sheet("DADOS")
    wsd.sheet_view.showGridLines = True
    _set_col_widths(wsd, {"A": 2, "B": 34, "C": 22, "D": 34, "E": 22, "F": 2})
    _merge_title(wsd, "B2:E2", "DADOS — NÃO EDITAR (BASE DOS GRÁFICOS)")
    wsd.row_dimensions[2].height = 22

    r = 4
    wsd[f"B{r}"] = "1) Ocorrências por mês (recorte final)"; wsd[f"B{r}"].font = Font(bold=True)
    r1s, _, r1e, _, _ = _add_table(wsd, r + 1, 2, df_mes, table_name="T_MES", style="TableStyleMedium9")

    r = r1e + 3
    wsd[f"B{r}"] = "2) Motivos (Top 12) — recorte final"; wsd[f"B{r}"].font = Font(bold=True)
    r2s, _, r2e, _, _ = _add_table(wsd, r + 1, 2, df_mot, table_name="T_MOT", style="TableStyleMedium9")

    r = r2e + 3
    wsd[f"B{r}"] = "3) Participação por Responsável (análise) — recorte final (Pizza)"; wsd[f"B{r}"].font = Font(bold=True)
    r3s, _, r3e, _, _ = _add_table(wsd, r + 1, 2, df_resp, table_name="T_RESP_PIE", style="TableStyleMedium9")

    r = r3e + 3
    wsd[f"B{r}"] = "4) Atrasadas por Responsável (análise) — conforme filtro (sem drill)"; wsd[f"B{r}"].font = Font(bold=True)
    r4s, _, r4e, _, _ = _add_table(wsd, r + 1, 2, df_atras, table_name="T_ATRAS_FILTRO", style="TableStyleMedium7")

    try:
        rng = f"C{r4s+1}:C{r4e}"
        wsd.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE"))
        )
    except Exception:
        pass

    _add_bar_chart_from_sheet(wsd, ws, "Ocorrências por mês (recorte)", 2, 3, r1s, r1e, "B12",
                              rotate_x_45=False, height=7.2, width=12.5, solid_fill_hex=BLUE)
    _add_bar_chart_from_sheet(wsd, ws, "Motivos (Top 12) — recorte", 2, 3, r2s, r2e, "D12",
                              rotate_x_45=True, height=7.2, width=12.5, solid_fill_hex=BLUE)
    _add_pie_chart_from_sheet(wsd, ws, "Participação por responsável (análise) — recorte", 2, 3, r3s, r3e, "B28",
                              height=7.2, width=12.5)
    _add_bar_chart_from_sheet(wsd, ws, "Atrasadas por responsável (análise) — conforme filtro", 2, 3, r4s, r4e, "D28",
                              rotate_x_45=True, height=7.2, width=12.5, solid_fill_hex=RED)

    ws2 = wb.create_sheet("RECORTE")
    ws2.sheet_view.showGridLines = True
    _merge_title(ws2, "A1:H1", "LISTA DE OCORRÊNCIAS — RECORTE FINAL (FILTRO + DRILL)")
    ws2.row_dimensions[1].height = 26

    cols_doc = [COL_CODIGO, COL_TITULO, COL_STATUS, COL_DATA, COL_CATEGORIA, COL_MOTIVO, COL_RESP_ANALISE, COL_SITUACAO]
    cols_doc = [c for c in cols_doc if c in dff.columns]

    dff_out = dff.copy()
    if COL_DATA in dff_out.columns:
        dff_out = dff_out.sort_values(COL_DATA, ascending=False).copy()
        dff_out[COL_DATA] = dff_out[COL_DATA].apply(br_date_str)
    if cols_doc:
        dff_out = dff_out[cols_doc].copy()

    _add_table(ws2, 3, 1, dff_out, table_name="T_RECORTE", style="TableStyleMedium9")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# =========================================================
# Persistência local (Streamlit Cloud / runtime)
# - Guarda a última planilha enviada e a última consulta.
# Observação: no Streamlit Cloud o disco é efêmero; em reinícios/redeploys os arquivos podem ser perdidos.
# =========================================================
DATA_STORE_DIR = os.path.join(os.path.dirname(__file__), "data_store")
LAST_META_PATH = os.path.join(DATA_STORE_DIR, "last_state.json")

def _ensure_data_store():
    os.makedirs(DATA_STORE_DIR, exist_ok=True)

def _now_iso():
    return pd.Timestamp.now(tz=None).strftime("%Y-%m-%d %H:%M:%S")

def save_uploaded_excel(file_bytes: bytes, original_name: str, sheet: str):
    _ensure_data_store()
    ts = pd.Timestamp.now(tz=None).strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", (original_name or "dados.xlsx"))
    saved_name = f"{ts}__{safe_name}"
    saved_path = os.path.join(DATA_STORE_DIR, saved_name)
    with open(saved_path, "wb") as f:
        f.write(file_bytes)

    meta = {
        "last_excel_path": saved_path,
        "original_name": original_name,
        "saved_name": saved_name,
        "loaded_at": _now_iso(),
        "sheet": sheet,
        "last_query": {},
    }
    with open(LAST_META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    return meta

def load_last_meta():
    try:
        with open(LAST_META_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def load_last_excel_bytes():
    meta = load_last_meta()
    if not meta:
        return None, None
    p = meta.get("last_excel_path")
    if not p or not os.path.exists(p):
        return None, meta
    with open(p, "rb") as f:
        return f.read(), meta

def save_last_query(state: dict):
    meta = load_last_meta() or {}
    meta["last_query"] = state or {}
    if "loaded_at" not in meta:
        meta["loaded_at"] = _now_iso()
    _ensure_data_store()
    with open(LAST_META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

def get_last_query():
    meta = load_last_meta() or {}
    return meta.get("last_query") or {}

def _df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "RECORTE") -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    out.seek(0)
    return out.read()

# =========================================================
# UI Streamlit
# =========================================================
st.set_page_config(page_title=APP_NAME, page_icon="📊", layout="wide")
require_login()
init_drill_state()

st.title(f"📊 {APP_NAME}")
st.caption("Painel interativo (Ocorrências) lado a lado com Motivos + Pizza + Atrasadas + Tabela por barra clicada.")

with st.sidebar:
    st.header("📥 Entrada")

    # Upload manual (prioridade)
    up = st.file_uploader("Envie o Excel (ex.: Consultas_RNC.xlsx)", type=["xlsx", "xlsm", "xls"])
    sheet = st.text_input("Nome da aba (sheet)", value=DEFAULT_SHEET)

    # Último arquivo salvo (se existir)
    last_bytes, last_meta = load_last_excel_bytes()
    use_last = False
    if (up is None) and (last_bytes is not None):
        use_last = st.checkbox("Usar automaticamente a última planilha salva", value=True)
        if use_last and last_meta:
            st.caption(f"Última planilha: **{last_meta.get('original_name','-')}** | carregada em **{last_meta.get('loaded_at','-')}**")

    st.divider()
    st.caption("Senha do app: QualidadeRS")

# Fonte de dados: upload > último salvo
excel_bytes = None
meta_loaded = None

if up is not None:
    excel_bytes = up.getvalue()
    meta_loaded = save_uploaded_excel(excel_bytes, getattr(up, "name", "dados.xlsx"), sheet)
elif (last_bytes is not None) and use_last:
    excel_bytes = last_bytes
    meta_loaded = last_meta

if excel_bytes is None:
    st.info("Envie o arquivo Excel para começar (ou habilite 'usar a última planilha salva' se já houver uma).")
    st.stop()

try:
    df_base = carregar_df(excel_bytes, sheet)
except Exception as e:
    st.error(f"Erro ao carregar: {e}")
    
# Cabeçalho: data dos dados / período do arquivo
p_min = br_date_str(df_base[COL_DATA].min()) if len(df_base) else "-"
p_max = br_date_str(df_base[COL_DATA].max()) if len(df_base) else "-"
if meta_loaded:
    st.info(f"📅 Dados: **{meta_loaded.get('original_name','-')}** | carregado em **{meta_loaded.get('loaded_at','-')}** | período no arquivo: **{p_min}** a **{p_max}**")
else:
    st.info(f"📅 Período no arquivo: **{p_min}** a **{p_max}**")

st.stop()

anos = sorted(df_base[COL_DATA].dt.year.dropna().unique().tolist())

last_q = get_last_query()

def _ss_init(k, v):
    if (k not in st.session_state) and (v is not None):
        st.session_state[k] = v

# restaura seleções principais (apenas 1x por sessão)
if "_restored_last_query" not in st.session_state:
    st.session_state["_restored_last_query"] = True
    _ss_init("ano_sel", last_q.get("ano_sel"))
    _ss_init("mes_sel", last_q.get("mes_sel"))
    _ss_init("resp_occ_sel", last_q.get("resp_occ_sel"))
    # drill + recorte
    for k in ["drill_level","drill_year","drill_month","table_focus_level","table_focus_value"]:
        _ss_init(k, last_q.get(k))
    # filtros por marcar
    _ss_init("filtros_sel", last_q.get("filtros_sel", {}))

c1, c2, c3, c4, c5 = st.columns([1, 1, 1.6, 1.2, 1.1])

with c1:
    ano_sel = st.selectbox("Ano", ["(Todos)"] + [str(a) for a in anos], index=0, key="ano_sel")
with c2:
    mes_sel = st.selectbox("Mês", ["(Todos)"] + [MESES_ABREV[m] for m in range(1, 13)], index=0, key="mes_sel")
with c3:
    if COL_RESP_OCORRENCIA in df_base.columns:
        resp_vals = sorted(df_base[COL_RESP_OCORRENCIA].dropna().astype(str).replace("nan", "").unique().tolist())
        resp_vals = [v for v in resp_vals if v != ""]
    else:
        resp_vals = []
    resp_occ_sel = st.selectbox("Resp. ocorrência", ["(Todos)"] + resp_vals, index=0, key="resp_occ_sel")
with c4:
    show_table = st.toggle("Mostrar tabela", value=True, key="show_table")
with c5:
    if st.button("🔄 Reset drill"):
        reset_drill()
        st.rerun()

with st.expander("Filtros por marcar (clique para abrir)", expanded=False):
    cols = st.columns(4)
    multi_filters = {}

    # filtros persistidos (última consulta)
    filtros_saved = st.session_state.get("filtros_sel", {}) or {}

    for i, col in enumerate(FILTROS_COLS):
        if col not in df_base.columns:
            continue
        vals = sorted(df_base[col].dropna().astype(str).replace("nan", "").unique().tolist())
        vals = [v for v in vals if v != ""]

        # default = último valor salvo (se ainda existir), senão todos
        default_vals = filtros_saved.get(col)
        if isinstance(default_vals, list):
            default_vals = [v for v in default_vals if v in vals]
        if not default_vals:
            default_vals = vals

        with cols[i % 4]:
            sel = st.multiselect(col, options=vals, default=default_vals, key=f"flt__{col}")
        multi_filters[col] = sel

    # guarda em memória para salvar no disco no final do run
    st.session_state["filtros_sel"] = {k: list(v) for k, v in multi_filters.items()}

df_filtrado = aplicar_filtros(df_base, ano_sel, mes_sel, resp_occ_sel, multi_filters)

total = int(len(df_filtrado))
situ = df_filtrado[COL_SITUACAO].apply(normalizar_situacao) if (COL_SITUACAO in df_filtrado.columns and total) else pd.Series([], dtype=str)
atras = int((situ == "ATRASADA").sum()) if total else 0
p_ini = br_date_str(df_filtrado[COL_DATA].min()) if total else "-"
p_fim = br_date_str(df_filtrado[COL_DATA].max()) if total else "-"

k1, k2, k3, k4 = st.columns(4)
k1.metric("Total ocorrências", total)
k2.metric("Em atraso (filtro)", atras)
k3.metric("Período", f"{p_ini} → {p_fim}")
k4.metric("Versão", APP_VERSION)

st.divider()
tab1, tab2 = st.tabs(["📈 Dashboard", "📦 Exportações (Excel/PDF)"])

with tab1:
    if not total:
        st.warning("Sem registros no filtro atual.")
        st.stop()

    # Ocorrências (dataset + figura)
    df_occ_plot, level_now, breadcrumb = occurrences_dataset(df_filtrado, ano_sel, mes_sel)
    fig_occ = fig_ocorrencias(df_occ_plot, level_now)

    # Base final (filtros + drill) para Motivos + Pizza
    df_final = apply_drill_filters(df_filtrado, ano_sel, mes_sel)
    df_mot_sel = calc_motivos(df_final, top_n=12)
    df_resp_sel = calc_resp_analise(df_final)
    df_atras_filtro = calc_atrasadas_por_filtro(df_filtrado)

    fig_mot = fig_motivos(df_mot_sel, "Motivos (Top 12) — seguindo seleção do gráfico Ocorrências")
    fig_pie = fig_pizza_participacao(df_resp_sel, "Participação por responsável (análise) — seleção do gráfico Ocorrências")
    titulo_ano = ano_sel if ano_sel != "(Todos)" else "Todos os anos"
    fig_atras = fig_atrasadas_vermelho(df_atras_filtro, f"Atrasadas por responsável (análise) — conforme filtro (Ano: {titulo_ano})")

    # Barra superior (controles drill/tabela)
    topbar1, topbar2, topbar3 = st.columns([1.2, 1.4, 3.4])
    with topbar1:
        if can_go_back(level_now, ano_sel, mes_sel):
            if st.button("⬅ Voltar (um nível)"):
                if go_back_one_level(level_now, ano_sel, mes_sel):
                    st.rerun()
    with topbar2:
        if st.button("🧹 Limpar seleção da tabela"):
            clear_table_focus()
            st.rerun()
    with topbar3:
        st.caption(f"📌 {breadcrumb}")

    # ✅ Linha 1: INTERATIVO (Ocorrências) lado a lado com Motivos
    colL, colR = st.columns(2)

    with colL:
        occ_event = None
        click_supported = True
        try:
            occ_event = st.plotly_chart(
                fig_occ,
                use_container_width=True,
                key="occ_chart",
                on_select="rerun",
                selection_mode="points",
            )
        except TypeError:
            click_supported = False
            st.plotly_chart(fig_occ, use_container_width=True)

        if click_supported:
            clicked = get_clicked_x(occ_event)
            if clicked is not None:
                # seleção para tabela
                st.session_state.table_focus_level = level_now
                st.session_state.table_focus_value = clicked

                # drill
                if level_now == "ANO" and ano_sel == "(Todos)":
                    try:
                        st.session_state.drill_year = int(clicked)
                        st.session_state.drill_level = "MES"
                        st.session_state.drill_month = None
                        st.rerun()
                    except Exception:
                        pass
                elif level_now == "MES" and mes_sel == "(Todos)":
                    mes_num = INV_MESES_ABREV.get(str(clicked))
                    if mes_num:
                        st.session_state.drill_month = int(mes_num)
                        st.session_state.drill_level = "SEMANA"
                        st.rerun()
                else:
                    st.rerun()

    with colR:
        st.plotly_chart(fig_mot, use_container_width=True)

    # Linha 2: Pizza + Atrasadas
    row2_left, row2_right = st.columns(2)
    with row2_left:
        st.plotly_chart(fig_pie, use_container_width=True)
    with row2_right:
        st.plotly_chart(fig_atras, use_container_width=True)

    # Tabela do recorte (barra clicada no gráfico interativo)
    if show_table:
        df_table_base = apply_drill_filters(df_filtrado, ano_sel, mes_sel)
        df_table = apply_table_focus(df_table_base)

        has_focus = (st.session_state.table_focus_level is not None) and (st.session_state.table_focus_value is not None)
        if not has_focus:
            st.info("Clique em uma barra do gráfico de Ocorrências para ver a lista completa das ocorrências daquele recorte.")
        else:
            info_sel = f"{st.session_state.table_focus_level}={st.session_state.table_focus_value}"
            with st.expander(f"📋 Ocorrências do recorte — {info_sel} (clique para abrir/fechar)", expanded=True):
                st.caption("Tabela com **todas as colunas** do recorte selecionado.")
                st.dataframe(df_table.sort_values(COL_DATA, ascending=False), use_container_width=True, height=420)

                # Download do recorte em Excel
                try:
                    excel_bytes = _df_to_excel_bytes(df_table.sort_values(COL_DATA, ascending=False), sheet_name="RECORTE")
                    st.download_button(
                        "⬇️ Baixar recorte em Excel",
                        data=excel_bytes,
                        file_name=f"recorte_{st.session_state.table_focus_level}_{st.session_state.table_focus_value}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error(f"Não foi possível gerar o Excel do recorte: {e}")

with tab2:
    if not total:
        st.info("Quando houver registros no filtro, as exportações ficam disponíveis.")
        st.stop()

    df_final_export = apply_drill_filters(df_filtrado, ano_sel, mes_sel)

    filtro_txt = _titulo_filtro(ano_sel, mes_sel, resp_occ_sel)
    drill_txt = []
    if ano_sel == "(Todos)" and st.session_state.drill_year is not None:
        drill_txt.append(f"Ano(clicado)={st.session_state.drill_year}")
    if mes_sel == "(Todos)" and st.session_state.drill_month is not None:
        drill_txt.append(f"Mês(clicado)={MESES_ABREV.get(int(st.session_state.drill_month), st.session_state.drill_month)}")
    if drill_txt:
        filtro_txt = filtro_txt + " | Drill: " + " ; ".join(drill_txt)

    total_final = int(len(df_final_export))
    situ_final = df_final_export[COL_SITUACAO].apply(normalizar_situacao) if (COL_SITUACAO in df_final_export.columns and total_final) else pd.Series([], dtype=str)
    atras_final = int((situ_final == "ATRASADA").sum()) if total_final else 0
    p_ini_final = br_date_str(df_final_export[COL_DATA].min()) if total_final else "-"
    p_fim_final = br_date_str(df_final_export[COL_DATA].max()) if total_final else "-"

    kpis_pdf = {"total": total_final, "atras": atras_final, "periodo": f"{p_ini_final} → {p_fim_final}"}

    st.subheader("📄 PDF do Dashboard (1 página, 4 gráficos)")
    try:
        df_occ_plot2, level_now2, _ = occurrences_dataset(df_filtrado, ano_sel, mes_sel)
        fig1 = fig_ocorrencias(df_occ_plot2, level_now2)

        df_mot_pdf = calc_motivos(df_final_export, top_n=12)
        df_resp_pdf = calc_resp_analise(df_final_export)
        df_atras_pdf = calc_atrasadas_por_filtro(df_filtrado)

        titulo_ano_pdf = ano_sel if ano_sel != "(Todos)" else "Todos os anos"
        fig2 = fig_motivos(df_mot_pdf, "Motivos (Top 12) — seleção do gráfico Ocorrências")
        fig3 = fig_pizza_participacao(df_resp_pdf, "Participação por responsável (análise) — seleção do gráfico Ocorrências")
        fig4 = fig_atrasadas_vermelho(df_atras_pdf, f"Atrasadas por responsável (análise) — conforme filtro (Ano: {titulo_ano_pdf})")

        pdf_bytes = build_dashboard_pdf_bytes(
            app_name=APP_NAME,
            filtro_txt=filtro_txt,
            kpis=kpis_pdf,
            figs_plotly=[fig1, fig2, fig3, fig4],
        )

        st.download_button(
            label="📄 Baixar PDF do Dashboard",
            data=pdf_bytes,
            file_name=f"Dashboard_{APP_NAME.replace(' ', '_')}.pdf",
            mime="application/pdf",
        )
    except Exception as e:
        st.error(f"Erro ao gerar PDF. Detalhe: {e}")
        st.caption("Se citar kaleido/Chrome, mantenha plotly==5.24.1 e kaleido==0.2.1 no requirements.txt")

    st.divider()
    st.subheader("📊 Resumo Excel (DASHBOARD + DADOS + RECORTE) — com Pizza")

    titulo_filtro = f"Reclamações — Filtro atual | {filtro_txt}"
    resumo_bytes = build_resumo_excel_bytes(df_final_export, df_filtrado, titulo_filtro)

    st.download_button(
        label="📥 Baixar Resumo Excel",
        data=resumo_bytes,
        file_name=f"Resumo_{APP_NAME.replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================================================
# Salva última consulta (persistência)
# =========================================================
try:
    save_last_query({
        "ano_sel": st.session_state.get("ano_sel"),
        "mes_sel": st.session_state.get("mes_sel"),
        "resp_occ_sel": st.session_state.get("resp_occ_sel"),
        "drill_level": st.session_state.get("drill_level"),
        "drill_year": st.session_state.get("drill_year"),
        "drill_month": st.session_state.get("drill_month"),
        "table_focus_level": st.session_state.get("table_focus_level"),
        "table_focus_value": st.session_state.get("table_focus_value"),
        "filtros_sel": st.session_state.get("filtros_sel", {}),
    })
except Exception:
    pass
